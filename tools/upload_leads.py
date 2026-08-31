#!/usr/bin/env python3
"""marketbase-upload-leads — bulk-ingest a CSV/XLSX into leads + lead_sources.

Behavior:
  - For each row, normalize the LinkedIn URL and UPSERT into `leads` keyed
    on linkedin_url. Existing rows have non-empty identity fields preserved
    (we only fill blanks).
  - Always insert a new row into `lead_sources` (append-only — every
    upload run leaves a footprint).
  - Optionally also append a `lead_qualifications` row when the file
    already carries a Persona Classification (so qualification history is
    captured at import time, not lost).
"""
from __future__ import annoacmens

import argparse
import csv
import json
import sys
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from lib import connect, normalize_linkedin_url, linkedin_urn, resolve_canonical_url


VALID_SOURCE_TYPES = {
    "buyer_monitor_likely_to_connect",
    "founder_network_export",
    "dripify_campaign_export",
    "dripify_reply_export",
    "linkedin_post_engagement",
    "manual_add",
    "apollo_export",
    "engagers_research",
    "find_senior_execs",
    "linkedin_people_search",
    "buyer_monitor_target",
}


# ── Column-detection helpers ─────────────────────────────────────────────────
URL_COL_CANDIDATES = [
    # Prefer vanity ("public") URLs when present. Dripify (and similar) exports
    # include BOTH a URN-encoded `Linkedin url` (https://.../in/ACoAAA…) and a
    # vanity `Linkedin public url` (https://.../in/firstname-lastname-XYZ).
    # The URN form makes leads unjoinable to engagement/post data later — see
    # the Maze re-ingest incident.
    "Linkedin public url", "LinkedIn Public URL", "linkedin_public_url",
    "LinkedIn URL", "linkedin_url", "Linkedin url", "Linkedin URL",
    "Connection LinkedIn URL", "Profile URL",
    "profile_url", "LinkedIn Profile URL", "linkedin",
]
NAME_COL_CANDIDATES = [
    "Name", "name", "Full Name", "Connection Name", "full_name",
]
FIRST_NAME_COLS = ["First name", "First Name", "first_name"]
LAST_NAME_COLS  = ["Last name", "Last Name", "last_name"]
HEADLINE_COLS   = ["Headline", "headline", "Connection Headline"]
TITLE_COLS      = ["Current Job Title", "Title", "title", "job_title", "Position", "position"]
COMPANY_COLS    = ["Current Company", "Company", "company", "Connection Company"]
COMPANY_URL_COLS = ["Company LinkedIn URL", "Company URL", "company_linkedin_url"]
CITY_COLS       = ["City", "city"]
COUNTRY_COLS    = ["Country", "country"]
LOCATION_COLS   = ["Location", "location"]
PERSONA_COLS    = ["Persona Classification", "persona", "Persona"]


def pick_col(header: list[str], candidates: list[str]) -> str | None:
    """Returns the first column name in `header` (case-insensitive) that
    matches any of the given candidates."""
    hdr_lower = {h.lower().lstrip("﻿"): h for h in header if h}
    for cand in candidates:
        h = hdr_lower.get(cand.lower())
        if h: return h
    return None


# ── Readers ──────────────────────────────────────────────────────────────────
def read_csv(path: Path):
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        rdr = csv.DictReader(f)
        for row in rdr:
            # strip BOMs in field names
            yield {k.lstrip("﻿"): v for k, v in row.items()}


def read_xlsx(path: Path, sheet: str | None = None):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    headers = [c.value for c in ws[1]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all((v is None or (isinstance(v, str) and not v.strip())) for v in r):
            continue
        yield {h: r[i] for i, h in enumerate(headers) if h is not None}


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(prog="marketbase-upload-leads")
    ap.add_argument("--client", required=True)
    ap.add_argument("--csv", help="Path to CSV input.")
    ap.add_argument("--xlsx", help="Path to XLSX input.")
    ap.add_argument("--sheet", help="XLSX sheet name (default: active).")
    ap.add_argument("--source-type", required=True, choices=sorted(VALID_SOURCE_TYPES),
                    help="Provenance tag.")
    ap.add_argument("--source-label", required=True,
                    help="Free text: the filename / search query / export run id.")
    ap.add_argument("--source-date", default=date.today().isoformat(),
                    help="YYYY-MM-DD (default: today).")
    ap.add_argument("--with-existing-qualification", action="store_true",
                    help="If the file has a 'Persona Classification' column, "
                         "ALSO write a lead_qualifications row using --qualifier-name + "
                         "--qualifier-version.")
    ap.add_argument("--qualifier-name", default="pre-marketbase-import",
                    help="Used when --with-existing-qualification is set.")
    ap.add_argument("--qualifier-version", default="legacy-1",
                    help="Used when --with-existing-qualification is set.")
    args = ap.parse_args()

    if not (args.csv or args.xlsx):
        ap.error("Provide --csv OR --xlsx.")

    src_path = Path(args.csv or args.xlsx)
    if not src_path.exists():
        ap.error(f"input not found: {src_path}")

    reader = read_csv(src_path) if args.csv else read_xlsx(src_path, args.sheet)

    src_date = datetime.fromisoformat(args.source_date).date() if "-" in args.source_date else None

    inserted = updated = sources_added = quals_added = skipped = 0
    seen_urls = set()

    with connect(args.client) as conn:
        with conn.cursor() as cur:
            for row in reader:
                # First-row column detection
                header = list(row.keys())
                url_col      = pick_col(header, URL_COL_CANDIDATES)
                name_col     = pick_col(header, NAME_COL_CANDIDATES)
                fn_col       = pick_col(header, FIRST_NAME_COLS)
                ln_col       = pick_col(header, LAST_NAME_COLS)
                hl_col       = pick_col(header, HEADLINE_COLS)
                title_col    = pick_col(header, TITLE_COLS)
                company_col  = pick_col(header, COMPANY_COLS)
                cu_col       = pick_col(header, COMPANY_URL_COLS)
                city_col     = pick_col(header, CITY_COLS)
                country_col  = pick_col(header, COUNTRY_COLS)
                loc_col      = pick_col(header, LOCATION_COLS)
                persona_col  = pick_col(header, PERSONA_COLS)
                if not url_col:
                    sys.exit(f"No LinkedIn URL column found. Headers: {header}")

                # process this row
                url = normalize_linkedin_url(row.get(url_col))
                if not url:
                    skipped += 1; continue
                # Resolve to the canonical person row (by member URN / vanity
                # slug) so a vanity-vs-URN URL difference updates the existing
                # lead instead of creating a duplicate.
                url = resolve_canonical_url(cur, url, urn_hint=linkedin_urn(url))
                if url in seen_urls:
                    # de-dup within file
                    pass
                seen_urls.add(url)

                # Compose identity fields from the row
                full_name = (row.get(name_col) if name_col else None) or \
                            " ".join(filter(None, [row.get(fn_col) if fn_col else None,
                                                   row.get(ln_col) if ln_col else None]))
                identity = {
                    "linkedin_url":        url,
                    "linkedin_urn":        linkedin_urn(url),
                    "name":                (full_name or "").strip() or None,
                    "headline":            (row.get(hl_col) or "").strip() or None,
                    "current_title":       (row.get(title_col) or "").strip() or None,
                    "current_company":     (row.get(company_col) or "").strip() or None,
                    "current_company_url": (row.get(cu_col) or "").strip() or None,
                    "city":                (row.get(city_col) or "").strip() or None,
                    "country":             (row.get(country_col) or "").strip() or None,
                }
                # If we have Location but no city/country, store in city
                if not identity["city"] and loc_col:
                    identity["city"] = (row.get(loc_col) or "").strip() or None

                # UPSERT lead — only fill blanks on conflict
                cur.execute("""
                    INSERT INTO leads (linkedin_url, linkedin_urn, name, headline,
                                       current_title, current_company, current_company_url,
                                       city, country)
                    VALUES (%(linkedin_url)s, %(linkedin_urn)s, %(name)s, %(headline)s,
                            %(current_title)s, %(current_company)s, %(current_company_url)s,
                            %(city)s, %(country)s)
                    ON CONFLICT (linkedin_url) DO UPDATE SET
                        name                = COALESCE(leads.name,                EXCLUDED.name),
                        headline            = COALESCE(leads.headline,            EXCLUDED.headline),
                        current_title       = COALESCE(leads.current_title,       EXCLUDED.current_title),
                        current_company     = COALESCE(leads.current_company,     EXCLUDED.current_company),
                        current_company_url = COALESCE(leads.current_company_url, EXCLUDED.current_company_url),
                        city                = COALESCE(leads.city,                EXCLUDED.city),
                        country             = COALESCE(leads.country,             EXCLUDED.country),
                        linkedin_urn        = COALESCE(leads.linkedin_urn,        EXCLUDED.linkedin_urn),
                        updated_at          = now()
                    RETURNING id, (xmax = 0) AS inserted
                """, identity)
                lead_id, was_inserted = cur.fetchone()
                if was_inserted: inserted += 1
                else: updated += 1

                # Always insert a lead_sources row (append-only)
                cur.execute("""
                    INSERT INTO lead_sources
                      (lead_id, source_type, source_label, source_date, raw_data)
                    VALUES (%s, %s, %s, %s, %s::jsonb)
                """, (lead_id, args.source_type, args.source_label, src_date,
                      json.dumps({k: (v.isoformat() if isinstance(v, (date, datetime)) else v)
                                  for k, v in row.items()}, default=str)))
                sources_added += 1

                # Optional: capture pre-existing persona classification
                if args.with_existing_qualification and persona_col:
                    persona = (row.get(persona_col) or "").strip()
                    if persona:
                        cur.execute("""
                            INSERT INTO lead_qualifications
                              (lead_id, qualifier_name, qualifier_version,
                               qualified, persona, reason, full_result)
                            VALUES (%s, %s, %s, %s, %s, %s, %s::jsonb)
                        """, (lead_id, args.qualifier_name, args.qualifier_version,
                              True, persona, "imported_qualification",
                              json.dumps({"source_persona": persona,
                                          "imported_from": args.source_label})))
                        quals_added += 1

        conn.commit()

    print(f"✓ upload complete")
    print(f"  source: {args.source_type!r} / {args.source_label!r}")
    print(f"  inserted leads: {inserted}")
    print(f"  updated leads:  {updated}")
    print(f"  lead_sources rows added: {sources_added}")
    print(f"  qualifications imported: {quals_added}")
    print(f"  rows skipped (no URL):   {skipped}")


if __name__ == "__main__":
    main()
