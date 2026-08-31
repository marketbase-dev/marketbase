#!/usr/bin/env python3
"""marketbase-stage-to-campaign — assign leads to a campaign.

Two ways to specify which leads:
  --from-csv / --from-xlsx <path>     : match by LinkedIn URL column
  --where-sql <fragment>              : raw SQL WHERE clause (e.g. 'persona = %(persona)s')
                                        with named params via --param key=value ...

Creates campaign (and the row in campaigns) if it doesn't exist.
UPSERTs campaign_members (lead_id, campaign_id) — never duplicates.
If a member already exists with a DIFFERENT status, leaves the status alone
unless --force-status is passed.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from lib import connect, normalize_linkedin_url


VALID_STATUSES = ("staged","uploaded","connection_requested","connection_accepted",
                  "message_sent","replied","meeting_booked","completed",
                  "removed_blocked","removed_other")

VALID_CHANNELS = ("linkedin_dm","linkedin_invite","email","dripify","smartlead",
                  "manual","other")


# ── Campaign naming convention ──────────────────────────────────────────────
# Required shape:
#   <source>_<persona>_<sequencer>_<descriptor>_<period>
#
# - 5 camelCase tokens separated by underscores
# - Last token is a 6-digit YYYYMM (e.g. 202606)
# - Each preceding token is camelCase, starts with a lowercase letter, no
#   underscores within tokens. Dots ARE allowed inside a token so version
#   suffixes read naturally (e.g. ...PersonasV5.4..., matching qualifier v5.4.0).
#
# Example (the canonical reference, used in this skill's docs):
#   likelyToAccept_allCloudAndSecPersonasV1_swanLinkedin_likelyToAcceptGeneric7mMessage_202606
#                ↑                       ↑              ↑                                  ↑
#            source                  persona       sequencer                       descriptor      period
#
# Token meanings:
#   source     — where the leads came from (likelyToAccept, northwindEngagers,
#                founderNetwork, dripifyTargets, …)
#   persona    — the persona / ICP slice we're targeting (e.g. allCloudAndSecPersonasV1,
#                cisoNonVendorV1, vpSecurityV2 — bump the suffix when the
#                criteria change so old campaigns keep their distinct name)
#   sequencer  — the outreach tool / channel handle. Often "<senderHandle><Tool>"
#                (e.g. swanLinkedin, nimoSmartlead). For purely manual = "manual".
#   descriptor — what's distinctive about THIS campaign: message variant,
#                throttle, follow-up cadence (e.g. likelyToAcceptGeneric7mMessage,
#                strictCloudSec3msgFollowup).
#   period     — YYYYMM, the campaign's intended SEND month, not the data-export
#                month. So leads scraped in 2026-05 sent in 2026-06 → 202606.
CAMPAIGN_NAME_PATTERN = re.compile(
    r"^[a-z][a-zA-Z0-9.]*"     # 1. source (camelCase; dots allowed for version suffixes)
    r"_[a-z][a-zA-Z0-9.]*"     # 2. persona
    r"_[a-z][a-zA-Z0-9.]*"     # 3. sequencer
    r"_[a-z][a-zA-Z0-9.]*"     # 4. descriptor
    r"_\d{6}$"                  # 5. period (YYYYMM)
)


def validate_campaign_name(name: str, strict: bool = False) -> tuple[bool, str]:
    """Returns (is_valid, message). On `strict=True`, sys.exits if invalid."""
    if CAMPAIGN_NAME_PATTERN.match(name):
        return True, ""
    msg = (
        f"⚠ campaign name {name!r} does not match the convention:\n"
        f"   <source>_<persona>_<sequencer>_<descriptor>_<period>\n"
        f"   5 camelCase tokens, last token is YYYYMM (6 digits)\n"
        f"   Example: likelyToAccept_allCloudAndSecPersonasV1_swanLinkedin_likelyToAcceptGeneric7mMessage_202606"
    )
    if strict:
        sys.exit(msg)
    return False, msg


def ensure_campaign(cur, name, channel, owner=None, persona=None, period=None,
                    description=None, sender_account=None, sender_email=None,
                    status="active", allow_nonstandard_name=False):
    ok, msg = validate_campaign_name(name)
    if not ok:
        if allow_nonstandard_name:
            sys.stderr.write(msg + "\n  (continuing because --allow-nonstandard-name)\n")
        else:
            sys.exit(msg + "\n  Pass --allow-nonstandard-name to override.")

    cur.execute("SELECT id FROM campaigns WHERE name = %s", (name,))
    row = cur.fetchone()
    if row: return row[0]
    cur.execute("""
        INSERT INTO campaigns
            (name, description, channel, status, owner, persona_target,
             period, sender_account, sender_email)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        RETURNING id
    """, (name, description, channel, status, owner, persona, period,
          sender_account, sender_email))
    return cur.fetchone()[0]


def read_urls_from_file(path: Path) -> list[str]:
    if path.suffix.lower() == ".csv":
        with open(path, newline="", encoding="utf-8", errors="replace") as f:
            rdr = csv.DictReader(f)
            urls = []
            for row in rdr:
                row = {k.lstrip("﻿"): v for k, v in row.items()}
                # Try common variants
                for k in ("LinkedIn URL","Linkedin url","Connection LinkedIn URL",
                          "linkedin_url","Linkedin public url","Profile URL"):
                    if k in row and row[k]:
                        urls.append(row[k]); break
            return urls
    else:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True); ws = wb.active
        h = [c.value for c in ws[1]]
        # find URL column
        for cand in ("LinkedIn URL","linkedin_url","Linkedin url","Connection LinkedIn URL"):
            if cand in h: ic = h.index(cand); break
        else:
            sys.exit(f"No URL column in xlsx; headers: {h}")
        return [r[ic] for r in ws.iter_rows(min_row=2, values_only=True) if r[ic]]


def main():
    ap = argparse.ArgumentParser(prog="marketbase-stage-to-campaign")
    ap.add_argument("--client", required=True)
    ap.add_argument("--campaign", required=True, help="Campaign name (e.g. dana_ciso_dripify_2026_05)")
    ap.add_argument("--channel", required=True, choices=VALID_CHANNELS)
    ap.add_argument("--owner")
    ap.add_argument("--persona-target", help="CISO / Cloud Security / etc.")
    ap.add_argument("--period", help="2026-05 etc.")
    ap.add_argument("--description")
    ap.add_argument("--sender-account",
                    help="Handle of the account that will send outreach (e.g. a LinkedIn or email account).")
    ap.add_argument("--sender-email")
    ap.add_argument("--campaign-status", default="active",
                    choices=("draft","staged","active","paused","closed"))
    ap.add_argument("--allow-nonstandard-name", action="store_true",
                    help="Bypass the <source>_<persona>_<sequencer>_<descriptor>_<period> "
                         "naming convention. Use only when migrating legacy campaign names.")

    # Lead-selection mode
    ap.add_argument("--from-csv", help="CSV with a LinkedIn URL column.")
    ap.add_argument("--from-xlsx", help="XLSX with a LinkedIn URL column.")
    ap.add_argument("--where-sql", help="SQL WHERE clause selecting from leads "
                    "(e.g. \"persona = 'CISO'\" — joined with lead_current_qualification AS q).")
    ap.add_argument("--all-with-persona-match", action="store_true",
                    help="Shortcut for: latest qualification has qualified=true AND persona is set.")

    # Member status
    ap.add_argument("--status", default="staged", choices=VALID_STATUSES,
                    help="Status to set for newly-added members.")
    ap.add_argument("--force-status", action="store_true",
                    help="Update status even if the member already exists.")
    ap.add_argument("--allow-multi-campaign", action="store_true",
                    help="Bypass the one-campaign-per-lead rule. By default, leads "
                         "already in any OTHER active campaign are skipped.")
    ap.add_argument("--allow-deal-company-leads", action="store_true",
                    help="Bypass the active-deal guard and stage leads even if their "
                         "employer has an open/won HubSpot deal (v_leads_at_deal_company). "
                         "Off by default — we never stage into active-deal companies.")
    ap.add_argument("--force-sequenced", action="store_true",
                    help="Bypass the active-HubSpot-sequence guard and stage leads even if "
                         "they carry flag:in_active_hubspot_sequence (already enrolled in a "
                         "HubSpot Sales sequence). Off by default — we don't double-message "
                         "someone who's already in a live HubSpot sequence.")
    ap.add_argument("--source", default="marketbase-stage-to-campaign",
                    help="last_status_source tag.")
    ap.add_argument("--notes")
    args = ap.parse_args()

    if not (args.from_csv or args.from_xlsx or args.where_sql or args.all_with_persona_match):
        ap.error("Provide one of --from-csv / --from-xlsx / --where-sql / --all-with-persona-match")

    with connect(args.client) as conn:
        with conn.cursor() as cur:
            campaign_id = ensure_campaign(
                cur, args.campaign, args.channel, args.owner,
                args.persona_target, args.period, args.description,
                args.sender_account, args.sender_email, args.campaign_status,
                allow_nonstandard_name=args.allow_nonstandard_name)

            # Resolve lead_ids
            if args.from_csv or args.from_xlsx:
                p = Path(args.from_csv or args.from_xlsx)
                urls = [normalize_linkedin_url(u) for u in read_urls_from_file(p)]
                urls = [u for u in urls if u]
                if not urls:
                    sys.exit("No URLs read from file.")
                cur.execute(
                    "SELECT id FROM leads WHERE linkedin_url = ANY(%s)",
                    (urls,))
                lead_ids = [r[0] for r in cur.fetchall()]
                missing = len(urls) - len(lead_ids)
                if missing:
                    print(f"  ⚠ {missing} URLs in file not found in leads table "
                          f"(upload them first with marketbase-upload-leads)", flush=True)
            elif args.all_with_persona_match:
                cur.execute("""
                    SELECT lead_id FROM lead_current_qualification
                    WHERE qualified = true AND persona IS NOT NULL AND persona <> ''
                """)
                lead_ids = [r[0] for r in cur.fetchall()]
            else:
                cur.execute(f"""
                    SELECT l.id FROM leads l
                    LEFT JOIN lead_current_qualification q ON q.lead_id = l.id
                    WHERE {args.where_sql}
                """)
                lead_ids = [r[0] for r in cur.fetchall()]

            print(f"  Selected {len(lead_ids)} leads for campaign {args.campaign!r}",
                  flush=True)

            # One-campaign-per-lead guard. Skip leads already in any OTHER active
            # campaign. Bypass with --allow-multi-campaign.
            skipped_multi = 0
            if lead_ids and not args.allow_multi_campaign:
                cur.execute("""
                    SELECT DISTINCT cm.lead_id
                    FROM campaign_members cm
                    JOIN campaigns c ON c.id = cm.campaign_id
                    WHERE c.status = 'active'
                      AND c.id <> %s
                      AND cm.lead_id = ANY(%s)
                """, (campaign_id, lead_ids))
                blocked = {r[0] for r in cur.fetchall()}
                if blocked:
                    skipped_multi = len(blocked)
                    lead_ids = [lid for lid in lead_ids if lid not in blocked]
                    print(f"  ⚠ {skipped_multi} leads already in another active "
                          f"campaign — skipping (use --allow-multi-campaign to override)",
                          flush=True)

            # Active-deal guard. Never stage leads whose CURRENT employer has an
            # open or won HubSpot deal (or a deleted deal still pending review).
            # This mirrors policy_active_deal_dq, but enforced HERE at stage-time
            # so a deal synced AFTER the last policy run still can't slip a lead
            # into a campaign — the second half of the belt-and-suspenders.
            # Override with --allow-deal-company-leads.
            skipped_deal = []
            if lead_ids and not args.allow_deal_company_leads:
                cur.execute("""
                    SELECT lead_id, name, linkedin_url, is_customer,
                           has_deleted_deal, open_stage, deal_companies
                    FROM v_leads_at_deal_company
                    WHERE lead_id = ANY(%s)
                    ORDER BY is_customer DESC, name
                """, (lead_ids,))
                skipped_deal = cur.fetchall()
                if skipped_deal:
                    blocked_deal = {r[0] for r in skipped_deal}
                    lead_ids = [lid for lid in lead_ids if lid not in blocked_deal]
                    print(f"  ⛔ {len(skipped_deal)} lead(s) work at a company with an "
                          f"active/won HubSpot deal — NOT staged "
                          f"(use --allow-deal-company-leads to override):", flush=True)
                    MAXSHOW = 50
                    for r in skipped_deal[:MAXSHOW]:
                        _, name, url, is_cust, has_del, stage, companies = r
                        kind = "CUSTOMER (won)" if is_cust else (
                            f"open deal — {stage}" if stage else "open deal")
                        if has_del:
                            kind += " [deal DELETED in HubSpot — pending review]"
                        comp = ", ".join(c for c in (companies or []) if c)
                        print(f"     • {name or '(no name)'} — {kind}"
                              f"{(' @ ' + comp) if comp else ''}  {url or ''}", flush=True)
                    if len(skipped_deal) > MAXSHOW:
                        print(f"     … and {len(skipped_deal) - MAXSHOW} more "
                              f"(full list: SELECT * FROM v_leads_at_deal_company)", flush=True)

            # Active-HubSpot-sequence guard. Never stage a lead who is already
            # enrolled in a live HubSpot Sales sequence (flag:in_active_hubspot_sequence)
            # into a MarketBase outbound campaign — that would double-message them across
            # two systems. Override with --force-sequenced. See CONVENTIONS.md.
            skipped_sequenced = []
            if lead_ids and not args.force_sequenced:
                cur.execute("""
                    SELECT DISTINCT lt.lead_id, l.name, l.linkedin_url
                    FROM lead_tags lt
                    JOIN leads l ON l.id = lt.lead_id
                    WHERE lt.tag = 'flag:in_active_hubspot_sequence'
                      AND lt.lead_id = ANY(%s)
                    ORDER BY l.name
                """, (lead_ids,))
                skipped_sequenced = cur.fetchall()
                if skipped_sequenced:
                    blocked_seq = {r[0] for r in skipped_sequenced}
                    lead_ids = [lid for lid in lead_ids if lid not in blocked_seq]
                    print(f"  ⛔ {len(skipped_sequenced)} lead(s) already enrolled in an "
                          f"active HubSpot sequence — NOT staged "
                          f"(use --force-sequenced to override):", flush=True)
                    MAXSHOW = 50
                    for _, name, url in skipped_sequenced[:MAXSHOW]:
                        print(f"     • {name or '(no name)'}  {url or ''}", flush=True)
                    if len(skipped_sequenced) > MAXSHOW:
                        print(f"     … and {len(skipped_sequenced) - MAXSHOW} more", flush=True)

            inserted = updated = unchanged = 0
            for lid in lead_ids:
                if args.force_status:
                    cur.execute("""
                        INSERT INTO campaign_members
                            (lead_id, campaign_id, status, last_status_source, notes)
                        VALUES (%s, %s, %s, %s, %s)
                        ON CONFLICT (lead_id, campaign_id) DO UPDATE SET
                            status = EXCLUDED.status,
                            last_status_source = EXCLUDED.last_status_source,
                            notes = COALESCE(EXCLUDED.notes, campaign_members.notes)
                        RETURNING (xmax = 0) AS inserted
                    """, (lid, campaign_id, args.status, args.source, args.notes))
                else:
                    cur.execute("""
                        INSERT INTO campaign_members
                            (lead_id, campaign_id, status, last_status_source, notes)
                        VALUES (%s, %s, %s, %s, %s)
                        ON CONFLICT (lead_id, campaign_id) DO NOTHING
                        RETURNING (xmax = 0) AS inserted
                    """, (lid, campaign_id, args.status, args.source, args.notes))
                row = cur.fetchone()
                if row is None:
                    unchanged += 1
                elif row[0]:
                    inserted += 1
                else:
                    updated += 1
        conn.commit()

    print(f"✓ staged to {args.campaign!r}")
    print(f"  new members:    {inserted}")
    print(f"  updated status: {updated}")
    print(f"  no-op (existed): {unchanged}")
    if not args.allow_multi_campaign and skipped_multi:
        print(f"  skipped (in another active campaign): {skipped_multi}")
    if not args.force_sequenced and skipped_sequenced:
        print(f"  withheld (in active HubSpot sequence): {len(skipped_sequenced)}")
    if skipped_deal:
        print(f"  withheld (employer has active/won deal): {len(skipped_deal)}")
        print()
        print(f"  ⚠ {len(skipped_deal)} lead(s) were withheld because their employer "
              f"has an active or won deal in HubSpot.")
        print( "    Recommended next steps:")
        print( "      • Leave them out — this is the deal-protection guard working as intended.")
        print( "      • If you think a deal is stale or already closed-lost, confirm in HubSpot.")
        print( "        The other side's daily sync will update MarketBase (or have it re-synced),")
        print( "        then re-run this stage command — released companies stage normally.")
        print( "      • To stage these deliberately anyway, re-run with --allow-deal-company-leads.")


if __name__ == "__main__":
    main()
