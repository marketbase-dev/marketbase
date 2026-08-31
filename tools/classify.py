#!/usr/bin/env python3
"""marketbase-classify

Generic, spec-driven classifier runner. Loads a classifier spec from the
`processors` registry, evaluates it against a set of leads, writes a
`lead_qualifications` row for each, and manages the qual: tag lifecycle.

The runner is logic-agnostic — what it does is decided by the YAML spec
under the classifier's `logic` block:

  logic.type: rules        — Python-expression IF/THEN clauses (safe eval)
  logic.type: gpt-prompt   — Jinja-style prompt -> JSON response via OpenAI
  logic.type: sql          — SELECT that returns (qualified, persona, reason)
                             per lead. Runner provides the lead set via a
                             temp table named `target_leads(lead_id uuid)`.

Usage:
  python3 classify.py --client Acme-AI --processor is-senior \
    --lead-url https://www.linkedin.com/in/somebody/
  python3 classify.py --client Acme-AI --processor is-senior --lead-file leads.csv
  python3 classify.py --client Acme-AI --processor is-senior \
    --where-tag potential_thought_leader
  python3 classify.py --client Acme-AI --processor is-senior \
    --all-tagged-for-this-processor
  python3 classify.py --client Acme-AI --processor is-senior@1.0 --lead-url ...
  python3 classify.py --client Acme-AI --processor is-senior --lead-url ... --re-classify
  python3 classify.py --client Acme-AI --processor is-senior --where-tag X --dry-run
"""
from __future__ import annotations

import argparse
import ast
import csv
import json
import os
import re
import sys
import traceback
from pathlib import Path
from typing import Any

import yaml
from psycopg.types.json import Jsonb

sys.path.insert(0, str(Path(__file__).resolve().parent))
from lib import connect, normalize_linkedin_url, register_processor_from_yaml


URL_COLUMNS = ("linkedin_url", "profile_url", "url", "linkedin", "LinkedIn URL")

# Per CONVENTIONS.md the canonical name is `state:qualification_queued`, but
# legacy DBs may still have the unprefixed `qualification_queued`. Read both.
QUEUE_TAGS = ("state:qualification_queued", "qualification_queued")

TAG_QUALIFIED = "qual:qualified"
TAG_DISQUALIFIED = "qual:disqualified"
TAGGED_BY = "marketbase-classify"


YAML_SPEC = """
name: marketbase-classify-runner
version: '1.0'
processor_type: classifier

description: >
  Generic, spec-driven classifier runner. Loads a classifier spec from the
  processors registry, evaluates its logic block (rules | gpt-prompt | sql)
  against a set of leads, writes lead_qualifications rows, and manages the
  qual: tag lifecycle (qualified/disqualified + clears qualification_queued).

inputs:
  fields_consulted:
    - processors.yaml_spec.logic
    - processors.yaml_spec.inputs.fields_consulted
    - leads.* (columns named in fields_consulted)
    - lead_qualifications.full_result (when fields_consulted references 'from-qualification:<name>:<key>')
    - lead_tags (--where-tag / --all-tagged-for-this-process selection)

decision_rule: |
  This runner does not itself decide qualification — it delegates to the
  named qualifier's logic block:

    rules:  short, side-effect-free Python expressions over `lead`. Allowed
            ops: arithmetic, comparison, AND/OR/NOT, IN/NOT IN, len(),
            .lower(), .upper(), .strip(), .split(...). Disallowed: import,
            exec, dunder attributes, unknown methods.
    gpt-prompt: Jinja-style template ({{ lead.field }}) -> OpenAI chat
            completion -> JSON response. Schema: {qualified: bool,
            persona: str?, reason: str?, ...arbitrary keys persisted into
            full_result}.
    sql:    a SELECT that, given a temp table target_leads(lead_id uuid),
            returns (lead_id, qualified, persona, reason). Runner inserts
            the qualifications based on each returned row.

  Per-lead lifecycle:
    1) Skip if a (lead, qualifier_name, qualifier_version) qualification
       already exists, unless --re-qualify.
    2) Gather declared inputs from leads.* and any 'from-qualification:*'
       cross-qualifier reads.
    3) Run the logic; capture qualified/persona/reason + full_result.
    4) INSERT lead_qualifications with qualified_at = clock_timestamp() so
       multiple rows in the same transaction get distinct timestamps
       (works correctly with the lead_current_qualification view tiebreaker).
    5) Tag swap: remove any *qualification_queued tag; UPSERT qual:qualified
       or qual:disqualified.

outputs:
  writes_to_tables: [lead_qualifications, lead_tags]
  tags_applied: [qual:qualified, qual:disqualified]
  tags_removed: [state:qualification_queued, qualification_queued]

rule_changes: |
  v 1.0 (initial): support logic.type in {rules, gpt-prompt, sql}; safe-eval
  for rules; OpenAI chat completions with response_format=json_object for
  gpt-prompt; temp-table-driven SELECT for sql.
"""


# ────────────────────────── safe expression evaluator ──────────────────────

_ALLOWED_NODES = (
    ast.Expression,
    ast.BoolOp, ast.UnaryOp, ast.BinOp, ast.Compare,
    ast.And, ast.Or, ast.Not,
    ast.Add, ast.Sub, ast.Mult, ast.Div, ast.FloorDiv, ast.Mod, ast.Pow,
    ast.USub, ast.UAdd,
    ast.Eq, ast.NotEq, ast.Lt, ast.LtE, ast.Gt, ast.GtE,
    ast.In, ast.NotIn, ast.Is, ast.IsNot,
    ast.Constant,
    ast.Name, ast.Load, ast.Store,
    ast.List, ast.Tuple, ast.Set, ast.Dict,
    ast.Subscript, ast.Slice,
    ast.Call, ast.Attribute,
    ast.IfExp,
    ast.keyword,
    # Comprehensions and generator expressions — needed for idiomatic
    # `any(k in title for k in [...])` rule patterns.
    ast.GeneratorExp, ast.ListComp, ast.SetComp, ast.DictComp,
    ast.comprehension,
)

_SAFE_ATTRS = {
    "lower", "upper", "strip", "lstrip", "rstrip", "split", "rsplit",
    "startswith", "endswith", "replace", "count", "find", "rfind",
    "isdigit", "isalpha", "isspace", "title",
    "get", "keys", "values", "items",
}

_SAFE_BUILTINS = {
    "len": len, "min": min, "max": max, "sum": sum,
    "any": any, "all": all, "abs": abs, "round": round,
    "int": int, "float": float, "str": str, "bool": bool,
    "list": list, "set": set, "tuple": tuple, "dict": dict,
    "sorted": sorted, "reversed": reversed,
    "True": True, "False": False, "None": None,
    "isinstance": isinstance,
}


def _validate_ast(node: ast.AST) -> None:
    for sub in ast.walk(node):
        if not isinstance(sub, _ALLOWED_NODES):
            raise ValueError(f"Disallowed expression node: {type(sub).__name__}")
        if isinstance(sub, ast.Attribute) and sub.attr not in _SAFE_ATTRS:
            raise ValueError(f"Disallowed attribute access: .{sub.attr}")
        if isinstance(sub, ast.Name) and sub.id.startswith("__"):
            raise ValueError(f"Disallowed name: {sub.id}")
        if isinstance(sub, ast.Attribute) and sub.attr.startswith("__"):
            raise ValueError(f"Disallowed dunder attribute: {sub.attr}")


def safe_eval(expr: str, env: dict) -> Any:
    """Evaluate a small Python expression against env. Raises on any
    disallowed construct.

    Important: names from `env` (like `lead`) are placed into the GLOBALS
    dict, not locals. Python comprehensions / generator expressions create
    their own scope, and only globals (not the caller's locals) are visible
    inside that scope — so `any(k in lead['headline'] for k in [...])`
    needs `lead` to be global from the evaluator's perspective."""
    tree = ast.parse(expr, mode="eval")
    _validate_ast(tree)
    safe_globals = {"__builtins__": {}}
    safe_globals.update(_SAFE_BUILTINS)
    safe_globals.update(env)
    return eval(
        compile(tree, "<rule>", "eval"),
        safe_globals,
        {},
    )


# ────────────────────────── Jinja-lite renderer ────────────────────────────

_JINJA_RE = re.compile(
    r"\{\{\s*([a-zA-Z_][a-zA-Z0-9_]*(?:\.[a-zA-Z_][a-zA-Z0-9_]*|\[[^\]]+\])*)\s*\}\}"
)


def render_template(template: str, env: dict) -> str:
    def _resolve(path: str) -> str:
        parts = re.findall(r"[a-zA-Z_][a-zA-Z0-9_]*|\[[^\]]+\]", path)
        if not parts:
            return ""
        head = parts[0]
        if head not in env:
            return ""
        cur: Any = env[head]
        for p in parts[1:]:
            if cur is None:
                return ""
            if p.startswith("["):
                key = p[1:-1].strip().strip("'\"")
                if isinstance(cur, dict):
                    cur = cur.get(key)
                elif isinstance(cur, (list, tuple)):
                    try:
                        cur = cur[int(key)]
                    except (ValueError, IndexError):
                        return ""
                else:
                    return ""
            else:
                if isinstance(cur, dict):
                    cur = cur.get(p)
                else:
                    cur = getattr(cur, p, None)
        return "" if cur is None else str(cur)

    return _JINJA_RE.sub(lambda m: _resolve(m.group(1)), template)


# ───────────────────────── input file parsing ──────────────────────────────

def read_urls_from_file(path: Path) -> list:
    if path.suffix.lower() in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("openpyxl required for XLSX input. pip install openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h) if h is not None else "" for h in rows[0]]
        col_idx = None
        for cand in URL_COLUMNS:
            if cand in headers:
                col_idx = headers.index(cand); break
        if col_idx is None:
            sys.exit(f"No URL column found in {path}. Expected one of: {URL_COLUMNS}")
        return [str(r[col_idx]) for r in rows[1:] if r[col_idx]]
    else:
        with open(path, newline="") as f:
            reader = csv.DictReader(f)
            col = None
            for cand in URL_COLUMNS:
                if cand in (reader.fieldnames or []):
                    col = cand; break
            if col is None:
                sys.exit(f"No URL column found in {path}. Expected one of: {URL_COLUMNS}")
            return [row[col] for row in reader if row.get(col)]


# ───────────────────────── process spec loading ────────────────────────────

def parse_process_arg(s: str):
    if "@" in s:
        n, v = s.split("@", 1)
        return n.strip(), v.strip()
    return s.strip(), None


def load_qualifier_spec(cur, name: str, version):
    if version:
        cur.execute("""
            SELECT name, version, processor_type, yaml_spec,
                   superseded_by IS NOT NULL AS superseded
            FROM processors WHERE name = %s AND version = %s
        """, (name, version))
    else:
        cur.execute("""
            SELECT name, version, processor_type, yaml_spec,
                   superseded_by IS NOT NULL AS superseded
            FROM processors
            WHERE name = %s AND superseded_by IS NULL
            ORDER BY created_at DESC LIMIT 1
        """, (name,))
    row = cur.fetchone()
    if not row:
        sys.exit(f"No process found for name={name!r}"
                 f"{' version=' + repr(version) if version else ''}.")
    db_name, db_version, ptype, yaml_text, superseded = row
    if ptype != "classifier":
        sys.exit(f"Processor {db_name}@{db_version} has type={ptype!r}; "
                 f"marketbase-classify only runs classifiers.")
    if superseded and not version:
        print(f"⚠  selected {db_name}@{db_version} which is superseded; "
              f"use --process {db_name}@<newer-version>.", file=sys.stderr)
    try:
        spec = yaml.safe_load(yaml_text)
    except yaml.YAMLError as e:
        sys.exit(f"Failed to parse YAML spec for {db_name}@{db_version}: {e}")
    if not isinstance(spec, dict):
        sys.exit(f"YAML for {db_name}@{db_version} is not a mapping.")
    spec.setdefault("name", db_name)
    spec.setdefault("version", db_version)
    spec["_db_version"] = db_version
    return spec


# ───────────────────────── target lead resolution ──────────────────────────

def resolve_targets(cur, *, lead_url, lead_file, where_tag,
                    all_tagged_for_this_process):
    if lead_url:
        u = normalize_linkedin_url(lead_url)
        cur.execute("SELECT id, name, linkedin_url FROM leads WHERE linkedin_url = %s", (u,))
        return cur.fetchall()
    if lead_file:
        urls = [normalize_linkedin_url(u) for u in read_urls_from_file(Path(lead_file))]
        urls = [u for u in urls if u]
        if not urls:
            return []
        cur.execute("""
            SELECT id, name, linkedin_url FROM leads
            WHERE linkedin_url = ANY(%s)
        """, (urls,))
        return cur.fetchall()
    if where_tag:
        cur.execute("""
            SELECT l.id, l.name, l.linkedin_url
            FROM leads l
            JOIN lead_tags t ON t.lead_id = l.id AND t.tag = %s
            ORDER BY t.tagged_at
        """, (where_tag,))
        return cur.fetchall()
    if all_tagged_for_this_process:
        cur.execute("""
            SELECT l.id, l.name, l.linkedin_url, MIN(t.tagged_at) AS first_queued
            FROM leads l
            JOIN lead_tags t ON t.lead_id = l.id AND t.tag = ANY(%s)
            GROUP BY l.id, l.name, l.linkedin_url
            ORDER BY first_queued
        """, (list(QUEUE_TAGS),))
        return [(r[0], r[1], r[2]) for r in cur.fetchall()]
    return []


# ───────────────────────── input gathering ─────────────────────────────────

class PrerequisiteMissingError(Exception):
    """The lead is missing data the qualifier requires. Surfaced clearly so
    the orchestrator (not this runner) can queue the prerequisite work."""


def gather_inputs(cur, lead_row: dict, fields_consulted: list) -> dict:
    env = dict(lead_row)
    if not fields_consulted:
        return env
    for f in fields_consulted:
        if not isinstance(f, str):
            continue
        if f.startswith("from-qualification:"):
            try:
                _, qname, key = f.split(":", 2)
            except ValueError:
                raise ValueError(
                    f"Malformed fields_consulted entry: {f!r}. "
                    f"Expected 'from-qualification:<qualifier_name>:<key>'.")
            cur.execute("""
                SELECT full_result FROM lead_qualifications
                WHERE lead_id = %s AND qualifier_name = %s
                ORDER BY qualified_at DESC, id DESC LIMIT 1
            """, (lead_row["id"], qname))
            row = cur.fetchone()
            if row is None or row[0] is None:
                raise PrerequisiteMissingError(
                    f"prerequisite missing: lead lacks a {qname!r} "
                    f"qualification (needed for field {f!r}). "
                    f"Run that qualifier first."
                )
            fr = row[0]
            if not isinstance(fr, dict) or key not in fr:
                raise PrerequisiteMissingError(
                    f"prerequisite missing: {qname!r} qualification exists "
                    f"but its full_result has no key {key!r}."
                )
            env[f] = fr[key]
            nskey = qname.replace("-", "_")
            env.setdefault(nskey, {})
            env[nskey][key] = fr[key]
        elif f.startswith("from-signal:"):
            try:
                _, ename, key = f.split(":", 2)
            except ValueError:
                raise ValueError(
                    f"Malformed fields_consulted entry: {f!r}. "
                    f"Expected 'from-signal:<enricher_name>:<key>'.")
            cur.execute("""
                SELECT payload FROM lead_signals
                WHERE lead_id = %s AND enricher_name = %s
                ORDER BY enriched_at DESC, id DESC LIMIT 1
            """, (lead_row["id"], ename))
            row = cur.fetchone()
            if row is None or row[0] is None:
                raise PrerequisiteMissingError(
                    f"prerequisite missing: lead lacks a {ename!r} "
                    f"signal (needed for field {f!r}). "
                    f"Run that enricher first."
                )
            payload = row[0]
            if not isinstance(payload, dict) or key not in payload:
                raise PrerequisiteMissingError(
                    f"prerequisite missing: {ename!r} signal exists "
                    f"but its payload has no key {key!r}."
                )
            env[f] = payload[key]
            nskey = ename.replace("-", "_")
            env.setdefault(nskey, {})
            env[nskey][key] = payload[key]
    return env


# ───────────────────────── logic runners ───────────────────────────────────

def run_rules(logic: dict, env: dict) -> dict:
    """First matching rule wins."""
    rules = logic.get("rules") or []
    if not isinstance(rules, list):
        raise ValueError("logic.rules must be a list")
    for i, rule in enumerate(rules):
        if not isinstance(rule, dict):
            raise ValueError(f"rule #{i} is not a mapping")
        cond = rule.get("if")
        if cond is None:
            outcome = rule.get("then") or {}
        else:
            try:
                matched = bool(safe_eval(str(cond), {"lead": env}))
            except PrerequisiteMissingError:
                raise
            except Exception as e:
                raise ValueError(
                    f"rule #{i} 'if' failed to evaluate: {e}\n  expr: {cond}"
                )
            if not matched:
                continue
            outcome = rule.get("then") or {}
        if not isinstance(outcome, dict):
            raise ValueError(f"rule #{i} 'then' must be a mapping")
        resolved = {}
        for k, v in outcome.items():
            if isinstance(v, str) and v.startswith("=") and len(v) > 1:
                try:
                    resolved[k] = safe_eval(v[1:], {"lead": env})
                except Exception as ex:
                    raise ValueError(f"rule #{i} 'then.{k}' expr failed: {ex}")
            else:
                resolved[k] = v
        resolved.setdefault("matched_rule", i)
        return resolved
    default = logic.get("default") or {}
    if not isinstance(default, dict):
        raise ValueError("logic.default must be a mapping")
    out = dict(default)
    out.setdefault("qualified", False)
    out.setdefault("reason", "no rule matched")
    return out


def run_gpt(logic: dict, env: dict) -> dict:
    template = logic.get("prompt_template")
    if not template:
        raise ValueError("logic.prompt_template is required for type=gpt-prompt")
    model = logic.get("model") or "gpt-4o-mini"
    system = logic.get("system_prompt") or (
        "You are a careful classifier. Return ONLY a JSON object that "
        "matches the requested schema. No prose, no markdown fences."
    )
    # Wrap env so templates can address fields both ways:
    #   {{ headline }}       — flat (legacy)
    #   {{ lead.headline }}  — namespaced (matches the rules-engine `lead['…']`
    #                          syntax AND the docstring example in SKILL.md)
    prompt = render_template(template, {"lead": env, **env})
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise ValueError("OPENAI_API_KEY env var required for gpt-prompt qualifiers")
    try:
        from openai import OpenAI
    except ImportError:
        raise ValueError("openai package required for gpt-prompt qualifiers. pip install openai")
    client = OpenAI(api_key=api_key)
    resp = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": prompt},
        ],
        response_format={"type": "json_object"},
        temperature=logic.get("temperature", 0),
    )
    content = resp.choices[0].message.content or "{}"
    try:
        parsed = json.loads(content)
    except json.JSONDecodeError as e:
        raise ValueError(f"GPT returned non-JSON: {e}\n--- content ---\n{content}")
    if not isinstance(parsed, dict):
        raise ValueError(f"GPT JSON is not an object: {parsed!r}")
    parsed.setdefault("qualified", False)
    parsed["_model"] = model
    parsed["_prompt_chars"] = len(prompt)
    return parsed


def run_sql_set(cur, logic: dict, lead_ids: list) -> dict:
    query = logic.get("query")
    if not query:
        raise ValueError("logic.query is required for type=sql")
    cur.execute("CREATE TEMP TABLE IF NOT EXISTS target_leads (lead_id uuid) ON COMMIT DROP")
    cur.execute("DELETE FROM target_leads")
    cur.executemany("INSERT INTO target_leads (lead_id) VALUES (%s)",
                    [(lid,) for lid in lead_ids])
    cur.execute(query)
    cols = [d.name for d in cur.description]
    if "lead_id" not in cols or "qualified" not in cols:
        raise ValueError(
            f"logic.query must return columns including 'lead_id' and "
            f"'qualified'; got {cols}")
    out: dict = {}
    for row in cur.fetchall():
        rec = dict(zip(cols, row))
        lid = rec.pop("lead_id")
        out[lid] = rec
    return out


# ───────────────────────── DB write helpers ────────────────────────────────

def existing_qualification(cur, lead_id, qualifier_name, qualifier_version):
    cur.execute("""
        SELECT 1 FROM lead_qualifications
        WHERE lead_id = %s AND qualifier_name = %s AND qualifier_version = %s
        LIMIT 1
    """, (lead_id, qualifier_name, qualifier_version))
    return cur.fetchone() is not None


def insert_qualification(cur, lead_id, qualifier_name, qualifier_version,
                         qualified, persona, reason, full_result):
    cur.execute("""
        INSERT INTO lead_qualifications
          (lead_id, qualifier_name, qualifier_version, qualified,
           persona, reason, full_result, qualified_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, clock_timestamp())
    """, (lead_id, qualifier_name, qualifier_version, bool(qualified),
          persona, reason, Jsonb(full_result)))


def tag_swap_after_qualification(cur, lead_id, qualified, reason):
    cur.execute("DELETE FROM lead_tags WHERE lead_id = %s AND tag = ANY(%s)",
                (lead_id, list(QUEUE_TAGS)))
    new_tag = TAG_QUALIFIED if qualified else TAG_DISQUALIFIED
    opp_tag = TAG_DISQUALIFIED if qualified else TAG_QUALIFIED
    cur.execute("DELETE FROM lead_tags WHERE lead_id = %s AND tag = %s",
                (lead_id, opp_tag))
    cur.execute("""
        INSERT INTO lead_tags (lead_id, tag, notes, tagged_by)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (lead_id, tag) DO UPDATE
          SET notes = EXCLUDED.notes,
              tagged_by = EXCLUDED.tagged_by,
              tagged_at = now()
    """, (lead_id, new_tag, reason, TAGGED_BY))


# ───────────────────────── main per-lead loop ──────────────────────────────

def lead_to_dict(cur, lead_id):
    cur.execute("""
        SELECT id, linkedin_url, linkedin_urn, public_id, name, headline,
               current_title, current_company, current_company_url,
               city, country, bio, last_enriched_at
        FROM leads WHERE id = %s
    """, (lead_id,))
    row = cur.fetchone()
    if row is None:
        return {}
    cols = [d.name for d in cur.description]
    return dict(zip(cols, row))


def normalize_outcome(raw: dict):
    qualified = bool(raw.get("qualified", False))
    persona = raw.get("persona")
    if persona is not None:
        persona = str(persona)
    reason = raw.get("reason")
    if reason is not None:
        reason = str(reason)
    return qualified, persona, reason, raw


def _jsonable(v):
    if hasattr(v, "isoformat"):
        return v.isoformat()
    if hasattr(v, "hex") and not isinstance(v, (bytes, bytearray)):
        try:
            return str(v)
        except Exception:
            return repr(v)
    return v


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Run a registered classifier against a set of leads.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--client", required=True)
    parser.add_argument("--processor", "--process", dest="processor", required=True,
                        help="Classifier name or name@version. "
                             "Default version: latest non-superseded.")
    target = parser.add_mutually_exclusive_group(required=True)
    target.add_argument("--lead-url", help="A single LinkedIn URL.")
    target.add_argument("--lead-file", help="CSV/XLSX with a URL column.")
    target.add_argument("--where-tag", help="Every lead carrying this tag.")
    target.add_argument("--all-tagged-for-this-processor",
                        "--all-tagged-for-this-process",
                        dest="all_tagged_for_this_processor", action="store_true",
                        help="Every lead with state:qualification_queued "
                             "(or legacy qualification_queued).")
    parser.add_argument("--re-classify", "--re-qualify",
                        dest="re_classify", action="store_true",
                        help="Re-run even if a qualification row already "
                             "exists for (lead, name, version).")
    parser.add_argument("--dry-run", action="store_true",
                        help="Compute decisions; don't write qualifications or tags.")
    args = parser.parse_args()

    try:
        _, _, action = register_processor_from_yaml(
            args.client, YAML_SPEC, created_by="marketbase-classify"
        )
        if action == "inserted":
            print("✓ registered processor marketbase-classify-runner@1.0")
    except Exception as e:
        print(f"⚠ runner self-registration failed (continuing anyway): {e}",
              file=sys.stderr)

    qname, qversion = parse_process_arg(args.processor)

    with connect(args.client) as conn:
        with conn.cursor() as cur:
            spec = load_qualifier_spec(cur, qname, qversion)
        qualifier_name = str(spec["name"])
        qualifier_version = str(spec["_db_version"])
        logic = spec.get("logic") or {}
        if not isinstance(logic, dict) or "type" not in logic:
            sys.exit(
                f"Qualifier {qualifier_name}@{qualifier_version} has no "
                f"`logic.type` in its YAML spec. marketbase-qualify needs one of: "
                f"rules, gpt-prompt, sql.\n"
                f"(Older specs describe rules in prose under `decision_rule` — "
                f"human-readable but not executable. Add a structured `logic:` "
                f"block to the spec to use this runner.)"
            )
        logic_type = logic.get("type")
        fields_consulted = (spec.get("inputs") or {}).get("fields_consulted") or []

        with conn.cursor() as cur:
            targets = resolve_targets(
                cur,
                lead_url=args.lead_url,
                lead_file=args.lead_file,
                where_tag=args.where_tag,
                all_tagged_for_this_process=args.all_tagged_for_this_processor,
            )

        if not targets:
            print("No target leads to qualify.")
            return 0

        print(f"Qualifier:  {qualifier_name}@{qualifier_version}  "
              f"(logic.type={logic_type})")
        print(f"Targets:    {len(targets)} lead(s)")
        if args.dry_run:
            print("(dry-run — no writes)")
        print()

        counts = {"qualified": 0, "disqualified": 0, "skipped": 0, "errors": 0}

        if logic_type == "sql":
            eligible = []
            for (lid, name, url) in targets:
                with conn.cursor() as cur:
                    if not args.re_classify and existing_qualification(
                        cur, lid, qualifier_name, qualifier_version
                    ):
                        counts["skipped"] += 1
                        print(f"  — skip (already qualified)  "
                              f"{name or '(no name)'}  {url}")
                        continue
                eligible.append((lid, name, url))
            if eligible:
                with conn.cursor() as cur:
                    try:
                        results = run_sql_set(
                            cur, logic, [lid for (lid, _, _) in eligible]
                        )
                    except Exception as e:
                        traceback.print_exc()
                        sys.exit(f"SQL qualifier failed: {e}")
                for (lid, name, url) in eligible:
                    raw = results.get(lid)
                    if raw is None:
                        counts["errors"] += 1
                        print(f"  ⚠ no SQL row returned for "
                              f"{name or '(no name)'}  {url}")
                        continue
                    qualified, persona, reason, full_result = normalize_outcome(raw)
                    if not args.dry_run:
                        with conn.cursor() as cur:
                            insert_qualification(
                                cur, lid, qualifier_name, qualifier_version,
                                qualified, persona, reason, full_result
                            )
                            tag_swap_after_qualification(cur, lid, qualified, reason)
                    counts["qualified" if qualified else "disqualified"] += 1
                    mark = "✓ qualified" if qualified else "✗ disqualified"
                    persona_s = f"  persona='{persona}'" if persona else ""
                    reason_s = f"  reason='{reason}'" if reason else ""
                    print(f"  {mark}{persona_s}{reason_s}  "
                          f"{name or ''}  {url}")
                if not args.dry_run:
                    conn.commit()
        else:
            for (lid, name, url) in targets:
                try:
                    with conn.cursor() as cur:
                        if not args.re_classify and existing_qualification(
                            cur, lid, qualifier_name, qualifier_version
                        ):
                            counts["skipped"] += 1
                            print(f"  — skip (already qualified)  "
                                  f"{name or ''}  {url}")
                            continue
                        lead_row = lead_to_dict(cur, lid)
                        if not lead_row:
                            counts["errors"] += 1
                            print(f"  ⚠ lead row vanished for id={lid}")
                            continue
                        env = gather_inputs(cur, lead_row, fields_consulted)

                    if logic_type == "rules":
                        raw = run_rules(logic, env)
                    elif logic_type == "gpt-prompt":
                        raw = run_gpt(logic, env)
                    else:
                        sys.exit(
                            f"Unknown logic.type={logic_type!r}. "
                            f"Supported: rules, gpt-prompt, sql."
                        )

                    raw_with_inputs = dict(raw)
                    raw_with_inputs["_inputs"] = {
                        k: _jsonable(v) for k, v in env.items() if k != "id"
                    }
                    qualified, persona, reason, full_result = normalize_outcome(
                        raw_with_inputs
                    )

                    if not args.dry_run:
                        with conn.cursor() as cur:
                            insert_qualification(
                                cur, lid, qualifier_name, qualifier_version,
                                qualified, persona, reason, full_result
                            )
                            tag_swap_after_qualification(cur, lid, qualified, reason)
                        conn.commit()

                    counts["qualified" if qualified else "disqualified"] += 1
                    mark = "✓ qualified" if qualified else "✗ disqualified"
                    persona_s = f"  persona='{persona}'" if persona else ""
                    reason_s = f"  reason='{reason}'" if reason else ""
                    print(f"  {mark}{persona_s}{reason_s}  "
                          f"{name or ''}  {url}")

                except PrerequisiteMissingError as e:
                    counts["errors"] += 1
                    conn.rollback()
                    print(f"  ⚠ {e}  {name or ''}  {url}")
                except Exception as e:
                    counts["errors"] += 1
                    conn.rollback()
                    print(f"  ⚠ ERROR ({type(e).__name__}: {e})  "
                          f"{name or ''}  {url}")

    print()
    print(f"Summary: qualified={counts['qualified']}  "
          f"disqualified={counts['disqualified']}  "
          f"skipped={counts['skipped']}  errors={counts['errors']}")
    if args.dry_run:
        print("(dry-run — no writes made)")
    return 0 if counts["errors"] == 0 else 2


if __name__ == "__main__":
    sys.exit(main())
