#!/usr/bin/env python3
"""marketbase-sync-conversations

Lead-targeted poll of Unipile conversations into the client's MarketBase.

For each lead the user specifies, walks every active `outbound_operators`
row, locates the LinkedIn chat between that operator and the lead (by
matching `attendee_provider_id` to the lead's URN), and incrementally
upserts the chat's messages + reactions into
`lead_conversations` / `lead_messages` / `lead_message_reactions`.

Side-effects per the CONVENTIONS.md auto-tagging rules:
  * any new inbound message  → tag `they:replied`
                              + bump campaign_members.status='replied'
                                (with the same terminal-status guard used by
                                 they:accepted_calendar_invite)
  * any new inbound reaction → tag `they:reacted_to_our_message`

Lead-targeted only — never auto-creates an unknown lead from a chat.

Usage:
  python3 sync_conversations.py --client Acme \\
    --lead-url https://www.linkedin.com/in/ACoAA...
  python3 sync_conversations.py --client Acme --lead-file leads.csv
  python3 sync_conversations.py --client Acme --lead-tag qual:qualified
  python3 sync_conversations.py --client Acme \\
    --lead-url https://… --chat-id <provider_chat_id>   # skips discovery
"""
from __future__ import annoacmens

import argparse
import concurrent.futures
import csv
import json
import os
import re
import sys
import threading
import time
import urllib.parse
import urllib.request
import urllib.error
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from lib import (
    connect, load_client_env, normalize_linkedin_url, linkedin_urn,
)


URL_COLUMNS = ("linkedin_url", "profile_url", "url", "linkedin", "LinkedIn URL")

# LinkedIn member URN token, as it appears in URN-encoded `/in/AC…` URLs and in
# `leads.linkedin_urn` (bare or as a full `urn:li:fsd_profile:AC…` string).
# All members share the `AC` prefix, but the following char varies by encoding
# (ACoAA, ACwAA, ACEAA, …) — match the whole family, not just ACoAA.
_LI_URN_RE = re.compile(r"AC[A-Za-z0-9_\-]{15,}")


def li_urn_token(v: str | None) -> str | None:
    """Extract the bare LinkedIn member URN from a raw URL slug or stored urn."""
    if not v:
        return None
    m = _LI_URN_RE.search(v)
    return m.group(0) if m else None

# campaign_members.status values that should NOT be silently overwritten
# when a new inbound message arrives. Mirrors the meeting-booked guard.
REPLIED_STATUS_GUARD = (
    "replied", "meeting_booked", "disqualified", "completed",
    "removed_blocked", "removed_other",
)


# ── Unipile client ─────────────────────────────────────────────────────────

class Unipile:
    """Minimal Unipile REST client. Token + base url come from env."""

    def __init__(self, base_url: str, token: str):
        self.base = base_url.rstrip("/")
        self.token = token

    def _get(self, path: str, **params) -> dict:
        q = urllib.parse.urlencode({k: v for k, v in params.items() if v is not None})
        url = f"{self.base}{path}?{q}" if q else f"{self.base}{path}"
        req = urllib.request.Request(
            url, headers={"X-API-KEY": self.token, "accept": "application/json"},
        )
        for attempt in range(3):
            try:
                with urllib.request.urlopen(req, timeout=30) as r:
                    return json.loads(r.read())
            except urllib.error.HTTPError as e:
                if e.code in (502, 503, 504) and attempt < 2:
                    time.sleep(1.5 ** attempt)
                    continue
                raise

    def find_chat_by_attendee(self, account_id: str, urn: str) -> dict | None:
        """Paginate this account's chats; return the first chat whose
        attendee_provider_id matches the given URN. None if not found."""
        cursor = None
        while True:
            params = {"account_id": account_id, "limit": 100}
            if cursor:
                params["cursor"] = cursor
            r = self._get("/api/v1/chats", **params)
            for c in r.get("items", []):
                if c.get("attendee_provider_id") == urn:
                    return c
            cursor = r.get("cursor")
            if not cursor or not r.get("items"):
                return None

    def fetch_chat_meta(self, chat_id: str) -> dict | None:
        try:
            return self._get(f"/api/v1/chats/{chat_id}")
        except urllib.error.HTTPError:
            return None

    def iter_messages(self, chat_id: str):
        """Yield every message in the chat (paginates via cursor)."""
        cursor = None
        while True:
            params = {"limit": 100}
            if cursor:
                params["cursor"] = cursor
            r = self._get(f"/api/v1/chats/{chat_id}/messages", **params)
            for m in r.get("items", []):
                yield m
            cursor = r.get("cursor")
            if not cursor or not r.get("items"):
                return


# ── Input parsing ──────────────────────────────────────────────────────────

def read_urls_from_file(path: Path) -> list[str]:
    if path.suffix.lower() in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("openpyxl required for XLSX. pip install openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h) if h is not None else "" for h in rows[0]]
        col_idx = next((headers.index(c) for c in URL_COLUMNS if c in headers), None)
        if col_idx is None:
            sys.exit(f"No URL column found in {path}. Expected one of: {URL_COLUMNS}")
        return [str(r[col_idx]) for r in rows[1:] if r[col_idx]]
    else:
        with open(path, newline="") as f:
            reader = csv.DictReader(f)
            col = next((c for c in URL_COLUMNS if c in (reader.fieldnames or [])), None)
            if col is None:
                sys.exit(f"No URL column found in {path}. Expected one of: {URL_COLUMNS}")
            return [row[col] for row in reader if row.get(col)]


# ── Conversation upsert helpers ────────────────────────────────────────────

def upsert_conversation(cur, operator_id, lead_id, channel, chat) -> str:
    """INSERT-or-fetch the lead_conversations row. Returns the conversation id."""
    cur.execute("""
        INSERT INTO lead_conversations (operator_id, lead_id, channel, provider_chat_id)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (operator_id, lead_id, channel) DO UPDATE
          SET provider_chat_id = EXCLUDED.provider_chat_id
        RETURNING id
    """, (operator_id, lead_id, channel, chat["id"]))
    return cur.fetchone()[0]


def upsert_message(cur, conversation_id, msg) -> tuple[str, bool]:
    """Upsert one message. Returns (message_id, is_new)."""
    cur.execute("""
        INSERT INTO lead_messages (conversation_id, provider_message_id, direction, sent_at, body, raw)
        VALUES (%s, %s, %s, %s, %s, %s::jsonb)
        ON CONFLICT (conversation_id, provider_message_id) DO UPDATE
          SET body = EXCLUDED.body, raw = EXCLUDED.raw
        RETURNING id, (xmax = 0) AS is_new
    """, (
        conversation_id,
        msg["provider_id"],
        "outbound" if msg.get("is_sender") else "inbound",
        msg["timestamp"],
        msg.get("text"),
        json.dumps(msg),
    ))
    mid, is_new = cur.fetchone()
    return mid, is_new


def upsert_reactions(cur, message_id, msg) -> int:
    """Upsert any reactions on the message. Returns count of newly-inserted
    inbound reactions on outbound messages (the trigger for the tag)."""
    msg_is_outbound = bool(msg.get("is_sender"))
    new_inbound_on_outbound = 0
    for r in msg.get("reactions") or []:
        reactor_outbound = bool(r.get("is_sender"))
        cur.execute("""
            INSERT INTO lead_message_reactions
              (message_id, reactor_provider_id, reactor_direction, reaction_value)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (message_id, reactor_provider_id, reaction_value) DO NOTHING
            RETURNING id
        """, (
            message_id,
            r.get("sender_id") or "",
            "outbound" if reactor_outbound else "inbound",
            r.get("value") or "",
        ))
        inserted = cur.fetchone() is not None
        if inserted and msg_is_outbound and not reactor_outbound:
            new_inbound_on_outbound += 1
    return new_inbound_on_outbound


def recompute_conversation_stats(cur, conversation_id):
    cur.execute("""
        UPDATE lead_conversations c SET
          first_message_at = (SELECT MIN(sent_at) FROM lead_messages WHERE conversation_id = c.id),
          last_message_at  = (SELECT MAX(sent_at) FROM lead_messages WHERE conversation_id = c.id),
          message_count    = (SELECT COUNT(*)     FROM lead_messages WHERE conversation_id = c.id),
          reaction_count   = (SELECT COUNT(*)     FROM lead_message_reactions r
                                JOIN lead_messages m ON m.id = r.message_id
                              WHERE m.conversation_id = c.id),
          last_synced_at   = now(),
          last_sync_error  = NULL
        WHERE c.id = %s
    """, (conversation_id,))


# ── Side effects: tag + status bump ────────────────────────────────────────

def apply_replied_side_effects(cur, lead_id):
    """Fired when a new inbound message lands. Returns (cm_updated, tag_applied)."""
    cur.execute("""
        INSERT INTO lead_tags (lead_id, tag, tagged_by)
        VALUES (%s, 'they:replied', 'marketbase-sync-conversations')
        ON CONFLICT (lead_id, tag) DO NOTHING
        RETURNING id
    """, (lead_id,))
    tag_applied = cur.fetchone() is not None
    cur.execute(f"""
        UPDATE campaign_members
           SET status = 'replied',
               last_status_at = now(),
               last_status_source = 'marketbase-sync-conversations:they:replied'
         WHERE lead_id = %s
           AND status NOT IN {REPLIED_STATUS_GUARD}
    """, (lead_id,))
    return cur.rowcount, tag_applied


def apply_reaction_side_effects(cur, lead_id):
    """Fired when a new inbound reaction on one of our messages lands."""
    cur.execute("""
        INSERT INTO lead_tags (lead_id, tag, tagged_by)
        VALUES (%s, 'they:reacted_to_our_message', 'marketbase-sync-conversations')
        ON CONFLICT (lead_id, tag) DO NOTHING
        RETURNING id
    """, (lead_id,))
    return cur.fetchone() is not None


# ── Per-lead sync ──────────────────────────────────────────────────────────

def sync_lead(conn, unipile_by_account: dict, operators: list[dict],
              lead_url: str, explicit_chat_id: str | None) -> dict:
    """Sync conversations for ONE lead across ALL active operators.

    Returns a per-lead summary dict. Commits before returning (durability)."""
    summary = {"lead_url": lead_url, "operators_checked": 0, "conversations": 0,
               "new_messages": 0, "new_inbound_messages": 0,
               "new_inbound_reactions": 0, "tags_applied": [],
               "campaign_status_bumps": 0, "skipped": None}

    normed = normalize_linkedin_url(lead_url)
    extracted = linkedin_urn(normed)

    with conn.cursor() as cur:
        cur.execute("SELECT id, linkedin_urn FROM leads WHERE linkedin_url = %s", (normed,))
        row = cur.fetchone()
        if not row:
            summary["skipped"] = "lead not in MarketBase (no auto-create)"
            return summary
        lead_id, stored_urn = row
        # Unipile's attendee_provider_id is the member's URN token. The slug we
        # extracted from the URL only works if the URL itself was in URN form
        # (/in/AC…); for vanity URLs we fall back to leads.linkedin_urn (which
        # the upload / enrichment pipeline can populate with the real URN, bare
        # or as a full `urn:li:fsd_profile:AC…` string). Members span several
        # URN prefixes (ACoAA, ACwAA, …) — accept the whole family.
        urn = li_urn_token(extracted) or li_urn_token(stored_urn)
        if not urn:
            summary["skipped"] = "no LinkedIn URN available (vanity URL + leads.linkedin_urn not backfilled)"
            return summary

        for op in operators:
            summary["operators_checked"] += 1
            api = unipile_by_account[op["unipile_account_id"]]
            # Each operator/conversation is synced inside its own SAVEPOINT so a
            # single failing chat records its error and rolls back just itself —
            # it never aborts the whole lead's transaction (and, via main's
            # rollback backstop, never poisons later leads either).
            cur.execute("SAVEPOINT op_sync")
            try:
                # Locate or look up the chat.
                chat = None
                if explicit_chat_id:
                    chat = {"id": explicit_chat_id, "attendee_provider_id": urn}
                else:
                    # Cached: do we already have a lead_conversations row?
                    cur.execute("""
                        SELECT provider_chat_id FROM lead_conversations
                        WHERE operator_id=%s AND lead_id=%s AND channel='linkedin'
                    """, (op["id"], lead_id))
                    cached = cur.fetchone()
                    if cached:
                        chat = {"id": cached[0], "attendee_provider_id": urn}
                    else:
                        # First time — paginate the operator's chats to find the URN.
                        chat = api.find_chat_by_attendee(op["unipile_account_id"], urn)
                if not chat:
                    cur.execute("RELEASE SAVEPOINT op_sync")
                    continue

                conv_id = upsert_conversation(cur, op["id"], lead_id, "linkedin", chat)

                new_inbound_msgs = 0
                new_inbound_reacts = 0
                new_msgs = 0
                for m in api.iter_messages(chat["id"]):
                    if not m.get("provider_id") or not m.get("timestamp"):
                        continue
                    msg_id, is_new = upsert_message(cur, conv_id, m)
                    if is_new:
                        new_msgs += 1
                        if not m.get("is_sender"):
                            new_inbound_msgs += 1
                    new_inbound_reacts += upsert_reactions(cur, msg_id, m)

                recompute_conversation_stats(cur, conv_id)

                if new_inbound_msgs > 0:
                    cm_updated, tag_applied = apply_replied_side_effects(cur, lead_id)
                    if tag_applied: summary["tags_applied"].append("they:replied")
                    summary["campaign_status_bumps"] += cm_updated
                if new_inbound_reacts > 0:
                    if apply_reaction_side_effects(cur, lead_id):
                        summary["tags_applied"].append("they:reacted_to_our_message")

                # Only fold counters in once the whole operator succeeded.
                summary["conversations"] += 1
                summary["new_messages"] += new_msgs
                summary["new_inbound_messages"] += new_inbound_msgs
                summary["new_inbound_reactions"] += new_inbound_reacts
                cur.execute("RELEASE SAVEPOINT op_sync")
            except Exception as e:
                cur.execute("ROLLBACK TO SAVEPOINT op_sync")
                err = str(e)[:500]
                summary.setdefault("errors", []).append(f"{op['display_name']}: {err}")
                # Persist the failure on the conversation row (if one exists) so
                # sync status lives in the DB. Done in its own savepoint so a
                # write failure here can't re-abort the transaction.
                cur.execute("SAVEPOINT op_err")
                try:
                    cur.execute("""
                        UPDATE lead_conversations
                           SET last_sync_error=%s, last_synced_at=now()
                         WHERE operator_id=%s AND lead_id=%s AND channel='linkedin'
                    """, (err, op["id"], lead_id))
                    cur.execute("RELEASE SAVEPOINT op_err")
                except Exception:
                    cur.execute("ROLLBACK TO SAVEPOINT op_err")

        conn.commit()
    return summary


# ── Main ───────────────────────────────────────────────────────────────────

def load_operators(conn) -> list[dict]:
    with conn.cursor() as cur:
        cur.execute("""
            SELECT id, display_name, unipile_account_id, channel_identifier, channel
            FROM outbound_operators
            WHERE is_active = true AND channel = 'linkedin'
            ORDER BY added_at
        """)
        cols = [d.name for d in cur.description]
        return [dict(zip(cols, r)) for r in cur.fetchall()]


def make_unipile_clients(env: dict, operators: list[dict]) -> dict:
    """One Unipile client per distinct (token, base_url). Today: one shared
    token per client MarketBase (per memory: a single Unipile workspace per
    client). Keyed by unipile_account_id for fast lookup.

    Picks the CLIENT pair (UNIPILE_ACCESS_TOKEN + UNIPILE_BASE_SERVER) when
    present; doesn't mix with the global UNIPILE_BASE_URL (different shards)."""
    if env.get("UNIPILE_ACCESS_TOKEN") and env.get("UNIPILE_BASE_SERVER"):
        token = env["UNIPILE_ACCESS_TOKEN"]
        base_url = f"https://{env['UNIPILE_BASE_SERVER']}"
    else:
        token = env.get("UNIPILE_API_KEY")
        base_url = env.get("UNIPILE_BASE_URL")
    if not token or not base_url:
        sys.exit("UNIPILE_ACCESS_TOKEN / UNIPILE_BASE_SERVER missing from env")
    client = Unipile(base_url, token)
    return {op["unipile_account_id"]: client for op in operators}


def resolve_urls(args, conn) -> list[str]:
    urls = []
    if args.lead_url:
        urls.append(args.lead_url)
    if args.lead_file:
        urls.extend(read_urls_from_file(Path(args.lead_file)))
    if args.lead_tag:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT l.linkedin_url
                FROM leads l JOIN lead_tags t ON t.lead_id = l.id
                WHERE t.tag = %s
            """, (args.lead_tag,))
            urls.extend(r[0] for r in cur.fetchall())
    # Dedup, preserve order
    seen, out = set(), []
    for u in urls:
        if u and u not in seen:
            seen.add(u); out.append(u)
    return out


def main() -> int:
    p = argparse.ArgumentParser(description="Sync Unipile conversations into MarketBase.")
    p.add_argument("--client", required=True)
    g = p.add_argument_group("lead selection (one or more)")
    g.add_argument("--lead-url", help="A single LinkedIn URL.")
    g.add_argument("--lead-file", help="CSV/XLSX with a URL column.")
    g.add_argument("--lead-tag", help="Sync every lead carrying this tag.")
    p.add_argument("--chat-id", help="Skip discovery; assume this provider_chat_id for the single --lead-url.")
    p.add_argument("--workers", type=int, default=1,
                   help="Sync this many leads concurrently (default 1 = sequential). "
                        "Leads are independent (each commits on its own), so this is "
                        "safe; each worker holds its own DB connection. Unipile "
                        "tolerates concurrent reads — 6-10 is a good range.")
    args = p.parse_args()

    if args.chat_id and not args.lead_url:
        sys.exit("--chat-id requires --lead-url (one specific lead).")
    if not (args.lead_url or args.lead_file or args.lead_tag):
        sys.exit("Provide --lead-url, --lead-file, or --lead-tag.")

    env = load_client_env(args.client)
    with connect(args.client) as conn:
        operators = load_operators(conn)
        if not operators:
            sys.exit(f"No active outbound_operators in {args.client} MarketBase. "
                     f"Run marketbase-register-operator first.")
        clients = make_unipile_clients(env, operators)
        urls = resolve_urls(args, conn)
        if not urls:
            print("No leads to sync."); return 0

        # End the read transaction now. In the parallel path below the workers
        # use their OWN thread-local connections, so this shared conn would
        # otherwise sit idle-in-transaction for the whole fan-out and get killed
        # by Postgres' idle_in_transaction_session_timeout (the __exit__ COMMIT
        # then blows up). Committing here leaves it merely idle, which is safe.
        conn.commit()

        workers = max(1, args.workers)
        print(f"Syncing {len(urls)} lead(s) against {len(operators)} operator(s) "
              f"[workers={workers}]...")
        totals = {"conversations": 0, "new_messages": 0, "new_inbound_messages": 0,
                  "new_inbound_reactions": 0, "campaign_status_bumps": 0,
                  "skipped": 0}
        error_leads = 0
        total = len(urls)

        def report(n, url, s, err):
            """Print one lead's result line and fold its counters into totals.
            Called only from the main thread (sequential loop or as_completed)."""
            nonlocal error_leads
            if err is not None:
                print(f"  [{n}/{total}] ERROR {url}: {err}", flush=True)
                error_leads += 1
                return
            if s["skipped"]:
                print(f"  [{n}/{total}] SKIP {url}: {s['skipped']}", flush=True)
                totals["skipped"] += 1
                return
            tag_note = (" tags:" + ",".join(s["tags_applied"])) if s["tags_applied"] else ""
            err_note = ""
            if s.get("errors"):
                err_note = "  ERRORS: " + "; ".join(s["errors"])
                error_leads += 1
            print(f"  [{n}/{total}] {url}  convos={s['conversations']}  "
                  f"new_msgs={s['new_messages']} (inbound={s['new_inbound_messages']})  "
                  f"new_reacts={s['new_inbound_reactions']}  cm_bumps={s['campaign_status_bumps']}"
                  f"{tag_note}{err_note}", flush=True)
            for k in totals:
                if k in s and k != "skipped": totals[k] += s[k]

        if workers == 1:
            for i, url in enumerate(urls, 1):
                try:
                    s = sync_lead(conn, clients, operators, url, args.chat_id)
                except Exception as e:
                    # A lead that escaped sync_lead's per-operator savepoints leaves
                    # the transaction aborted — roll back so the next lead is usable.
                    conn.rollback()
                    report(i, url, None, e)
                    continue
                report(i, url, s, None)
        else:
            # Parallel path. Leads are independent (each sync_lead commits on its
            # own), so they fan out cleanly — but a psycopg connection is NOT
            # thread-safe, so every worker thread holds its OWN connection
            # (cached thread-local, reused across that thread's leads). The
            # Unipile client is stateless HTTP and safe to share. Futures are
            # consumed in the MAIN thread via as_completed, so report()/totals
            # need no lock. The shared `conn` is left for operator/url loading.
            _tl = threading.local()
            _conns: list = []
            _conns_lock = threading.Lock()

            def _worker_conn():
                c = getattr(_tl, "conn", None)
                if c is None:
                    c = connect(args.client)
                    _tl.conn = c
                    with _conns_lock:
                        _conns.append(c)
                return c

            def _work(n_url):
                n, url = n_url
                wconn = _worker_conn()
                try:
                    return (n, url, sync_lead(wconn, clients, operators, url, args.chat_id), None)
                except Exception as e:
                    wconn.rollback()
                    return (n, url, None, e)

            done = 0
            with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as ex:
                futs = [ex.submit(_work, (i, url)) for i, url in enumerate(urls, 1)]
                for fut in concurrent.futures.as_completed(futs):
                    _n, url, s, err = fut.result()
                    done += 1
                    report(done, url, s, err)
            for c in _conns:
                try:
                    c.close()
                except Exception:
                    pass

        print(f"\n=== Totals ===")
        for k, v in totals.items():
            print(f"  {k}: {v}")
        if error_leads:
            print(f"  leads_with_errors: {error_leads}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
