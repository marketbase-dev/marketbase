#!/usr/bin/env python3
"""marketbase-tag-lead

Apply (or remove) a tag on one or more leads.

Tags are mutable, multi-valued categorization stored in `lead_tags`. They
differ from `lead_sources` (immutable provenance) and `lead_qualifications`
(algorithmic classification history).

Usage:
  # Tag a single lead
  python3 tag_lead.py --client Acme-AI \
    --lead-url https://www.linkedin.com/in/jonathan-r-bland/ \
    --tag thought-leader --notes "Carousel 1 participant" --tagged-by claude

  # Tag many leads from a CSV/XLSX (column auto-detected: linkedin_url, profile_url, url)
  python3 tag_lead.py --client Acme-AI --lead-file leaders.csv --tag thought-leader

  # Remove a tag
  python3 tag_lead.py --client Acme-AI --lead-url https://… --tag old-status --remove
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from lib import connect, normalize_linkedin_url


URL_COLUMNS = ("linkedin_url", "profile_url", "url", "linkedin", "LinkedIn URL")

# Tags allowed to carry a `notes` value (per CONVENTIONS.md "Note allow-list").
# Value = (required: bool, expected shape).
NOTES_ALLOWLIST: dict[str, tuple[bool, str]] = {
    "they:redirected_to_colleague": (True,  "comma-separated colleague names"),
    "we:sent_referral_ask":         (True,  "comma-separated proposed names"),
    "plan:snooze":                  (False, "resume date YYYY-MM-DD"),
    "plan:deal_hold":               (True,  "Active deal at employer — <deal_company> (<open_stage>)"),
    "plan:verify_disqualification": (False, "one-phrase reason for the doubt"),
    "qual:disqualified":            (False, "one-phrase DQ reason (e.g. 'VAR/reseller', 'wants consulting fee', 'out-of-geo:India', 'met -> irrelevant', 'security vendor')"),
    "flag:parallel_threads":        (True,  "Active: <chat_id> (<sender>); Dormant: <chat_id> (<sender>); Reason: <why dormant>; Updated: YYYY-MM-DD"),
    "flag:manually_approved":       (True,  "<who> approved YYYY-MM-DD; overrides <what flag/auto-DQ>"),
    "flag:borderline_qualified":    (True,  "one-phrase reason for the doubt (e.g. job-seeker headline + ongoing-involvement scope Q)"),
    "flag:meeting_not_attended":    (True,  "<meeting date/time> — <cause> (e.g. 'Jun 11 2pm WIB — invite undelivered, landed in spam')"),
    "flag:possible_consultant":     (True,  "<firm> — consulting/integrator/MSP; possible consultant/partner not direct buyer; verify ICP fit"),
    "flag:security_vendor":         (True,  "<firm> — security product vendor; employee not a direct-buyer ICP"),
    "flag:awaiting_shimon_decision": (True, "<the decision Dana must make> — <one-line context>; clear once decided (replace with the resulting qual:/plan:)"),
}


# `they:*` tags that mean "a meeting is on the calendar" — applying any of them
# promotes campaign_members.status to 'meeting_booked' and clears the redundant
# `plan:book_meeting` tag. `they:accepted_calendar_invite` = prospect accepted an
# invite we sent; `they:booked_meeting` = prospect self-booked (e.g. Calendly).
# Both flip the status so WE record the booking ourselves rather than leaking it
# to Smartlead's reconciliation. (Convention updated 2026-06-24, alice: a self-booking
# now counts as booked — superseding the older Calendly-ghost carve-out.)
# Documented in CONVENTIONS.md.
MEETING_BOOKED_TAGS = frozenset({
    "they:accepted_calendar_invite",
    "they:booked_meeting",
})
# Campaign-member statuses preserved when a meeting-booked tag fires
# (don't silently revive a terminally-removed or already-completed lead).
MEETING_BOOKED_STATUS_GUARD = (
    "meeting_booked", "disqualified", "completed",
    "removed_blocked", "removed_other",
)

# `they:*` tags that mean "the lead sent us a reply" — applying any of them
# bumps campaign_members.status to 'replied' (with the same terminal-status
# guard pattern used for meeting_booked). Documented in CONVENTIONS.md.
REPLY_TAGS = frozenset({
    "they:replied",
    "they:willing_to_give_feedback",
    "they:declined_to_give_feedback",
    "they:politely_greeted",
    "they:redirected_to_colleague",
    "they:gave_deep_feedback",
    "they:skeptical_but_engaged",
    "they:asked_about_scope",
    "they:job_seeker",
    "they:confused",
})
REPLIED_STATUS_GUARD = (
    "replied", "meeting_booked", "disqualified", "completed",
    "removed_blocked", "removed_other",
)


def apply_replied_side_effect(cur, lead_id, tag: str) -> int:
    """Bump campaign_members.status='replied' for this lead, terminal-status
    guarded. Returns count of rows updated."""
    cur.execute(f"""
        UPDATE campaign_members
           SET status = 'replied',
               last_status_at = now(),
               last_status_source = 'marketbase-tag-lead:' || %s
         WHERE lead_id = %s
           AND status NOT IN {REPLIED_STATUS_GUARD}
    """, (tag, lead_id))
    return cur.rowcount


def apply_meeting_booked_side_effects(cur, lead_id, tag: str) -> tuple[int, int]:
    """When a meeting-booked tag (`they:accepted_calendar_invite` /
    `they:booked_meeting`) is applied to a lead, propagate to campaign_members +
    clean up the now-redundant `plan:book_meeting` tag.
    Returns (campaign_rows_updated, plan_tags_removed)."""
    cur.execute(f"""
        UPDATE campaign_members
           SET status = 'meeting_booked',
               last_status_at = now(),
               last_status_source = 'marketbase-tag-lead:' || %s
         WHERE lead_id = %s
           AND status NOT IN {MEETING_BOOKED_STATUS_GUARD}
    """, (tag, lead_id))
    cm_updated = cur.rowcount
    cur.execute(
        "DELETE FROM lead_tags WHERE lead_id=%s AND tag='plan:book_meeting'",
        (lead_id,),
    )
    plan_removed = cur.rowcount
    return cm_updated, plan_removed


def validate_notes_against_allowlist(tag: str, notes: str | None) -> None:
    """Enforce the CONVENTIONS.md note allow-list. Warn now; intended to
    hard-fail in a later release. Apply BEFORE writing — keeps stderr noise
    out of bulk-tag runs where notes are empty for all leads."""
    has_notes = notes is not None and notes != ""
    entry = NOTES_ALLOWLIST.get(tag)
    if entry is None:
        if has_notes:
            print(f"  ⚠ tag '{tag}' is not on the note allow-list — "
                  f"`lead_tags.notes` will be ignored in a future release. "
                  f"See ~/.claude/tools/MarketBase/CONVENTIONS.md", file=sys.stderr)
        return
    required, shape = entry
    if required and not has_notes:
        print(f"  ⚠ tag '{tag}' requires a note ({shape}). "
              f"Will be hard-failed in a future release.", file=sys.stderr)


def read_urls_from_file(path: Path) -> list[str]:
    """Auto-detect a URL column from CSV or XLSX."""
    if path.suffix.lower() in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("openpyxl required for XLSX input. pip install openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows: return []
        headers = [str(h) if h is not None else "" for h in rows[0]]
        col_idx = None
        for cand in URL_COLUMNS:
            if cand in headers:
                col_idx = headers.index(cand); break
        if col_idx is None:
            sys.exit(f"No URL column found in {path}. Expected one of: {URL_COLUMNS}")
        return [str(r[col_idx]) for r in rows[1:] if r[col_idx]]
    else:
        with open(path, newline="") as f:
            reader = csv.DictReader(f)
            col = None
            for cand in URL_COLUMNS:
                if cand in reader.fieldnames or []:
                    col = cand; break
            if col is None:
                sys.exit(f"No URL column found in {path}. Expected one of: {URL_COLUMNS}")
            return [row[col] for row in reader if row.get(col)]


def apply_tag(client: str, urls: list[str], tag: str, notes: str | None,
              tagged_by: str | None, remove: bool) -> dict[str, int]:
    """Returns counts: {'matched_leads', 'tagged', 'untagged', 'unknown_urls'}."""
    if not remove:
        validate_notes_against_allowlist(tag, notes)
    counts = {"matched_leads": 0, "tagged": 0, "untagged": 0, "unknown_urls": 0,
              "meeting_booked_promotions": 0, "plan_book_meeting_removed": 0,
              "replied_promotions": 0}
    normed = [normalize_linkedin_url(u) for u in urls if u]
    normed = [u for u in normed if u]

    with connect(client) as conn:
        with conn.cursor() as cur:
            for u in normed:
                cur.execute("SELECT id FROM leads WHERE linkedin_url = %s", (u,))
                row = cur.fetchone()
                if not row:
                    counts["unknown_urls"] += 1
                    print(f"  ⚠ no lead found for {u}")
                    continue
                lead_id = row[0]
                counts["matched_leads"] += 1
                if remove:
                    cur.execute("DELETE FROM lead_tags WHERE lead_id=%s AND tag=%s",
                                (lead_id, tag))
                    if cur.rowcount:
                        counts["untagged"] += 1
                else:
                    cur.execute("""
                        INSERT INTO lead_tags (lead_id, tag, notes, tagged_by)
                        VALUES (%s, %s, %s, %s)
                        ON CONFLICT (lead_id, tag) DO UPDATE
                          SET notes = EXCLUDED.notes,
                              tagged_by = EXCLUDED.tagged_by,
                              tagged_at = now()
                    """, (lead_id, tag, notes, tagged_by))
                    counts["tagged"] += 1
                    if tag in MEETING_BOOKED_TAGS:
                        cm_updated, plan_removed = apply_meeting_booked_side_effects(cur, lead_id, tag)
                        counts["meeting_booked_promotions"] += cm_updated
                        counts["plan_book_meeting_removed"] += plan_removed
                    elif tag in REPLY_TAGS:
                        counts["replied_promotions"] += apply_replied_side_effect(cur, lead_id, tag)
        conn.commit()
    return counts


def main() -> int:
    parser = argparse.ArgumentParser(description="Apply or remove a tag on one or more leads.")
    parser.add_argument("--client", required=True)
    parser.add_argument("--tag", required=True)
    parser.add_argument("--lead-url", help="A single LinkedIn URL to tag.")
    parser.add_argument("--lead-file", help="CSV/XLSX with a URL column.")
    parser.add_argument("--notes", default=None)
    parser.add_argument("--tagged-by", default=None,
                        help="Who/what is applying the tag (e.g. 'claude', 'alice').")
    parser.add_argument("--remove", action="store_true",
                        help="Remove the tag instead of applying it.")
    args = parser.parse_args()

    if not args.lead_url and not args.lead_file:
        sys.exit("--lead-url or --lead-file is required")

    urls = []
    if args.lead_url: urls.append(args.lead_url)
    if args.lead_file: urls.extend(read_urls_from_file(Path(args.lead_file)))

    counts = apply_tag(args.client, urls, args.tag, args.notes, args.tagged_by, args.remove)

    verb = "untagged" if args.remove else "tagged"
    print(f"\n{verb}: {counts[verb]}  matched: {counts['matched_leads']}  unknown: {counts['unknown_urls']}")
    if counts["meeting_booked_promotions"] or counts["plan_book_meeting_removed"]:
        print(f"  side-effects: campaign_members→meeting_booked: {counts['meeting_booked_promotions']}  "
              f"plan:book_meeting removed: {counts['plan_book_meeting_removed']}")
    if counts["replied_promotions"]:
        print(f"  side-effects: campaign_members→replied: {counts['replied_promotions']}")
    return 0 if counts["unknown_urls"] == 0 else 2


if __name__ == "__main__":
    sys.exit(main())
