#!/usr/bin/env python3
"""demand_gen_signals_enricher@1.0

For each target lead:
  1. Fetch their last ~50 LinkedIn posts via Saleleads /api/v1/user/posts
  2. Compute posting stats (3/6/12-month windows; original posts only)
  3. Run 5 GPT classification prompts (base + 4 repass) — gpt-4o-mini, temp=0
  4. Merge into a typed signal payload (numerics as int/float, not strings)
  5. Write to `lead_signals` table under (enricher_name=demand_gen_signals_enricher,
     enricher_version=1.0)

Faithful port of the Acme-AI carousel-1 ETL Node.js scripts
(<your-local-tools-dir>/*.js).
Reuses the prompt text + truncation rules verbatim so outputs stay
calibration-compatible with the legacy 388 + 850 lead values already in
lead_signals @legacy-2026-05.

CLI:
  python3 demand_gen_signals_enricher.py --client Acme-AI \\
      --lead-url https://www.linkedin.com/in/somebody/

  python3 demand_gen_signals_enricher.py --client Acme-AI \\
      --lead-file leads.csv [--refresh] [--limit 50]

  python3 demand_gen_signals_enricher.py --client Acme-AI \\
      --where-tag potential_thought_leader [--refresh] [--limit 50]

Skips leads that already have a current-version signal row unless
--refresh. Idempotent — ON CONFLICT DO UPDATE on (lead_id, enricher_name,
enricher_version).
"""
from __future__ import annoacmens

import argparse
import csv
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from psycopg.types.json import Jsonb

sys.path.insert(0, str(Path(__file__).resolve().parent))
from lib import connect, normalize_linkedin_url, register_processor_from_yaml
from demand_gen_prompts import (
    PROMPTS, MODEL, TEMPERATURE, RESPONSE_FORMAT, compute_stats,
)


ENRICHER_NAME = "demand_gen_signals_enricher"
ENRICHER_VERSION = "1.1"

SL_HOST = "fresh-linkedin-scraper-api.p.rapidapi.com"
SL_BASE = f"https://{SL_HOST}"


YAML_SPEC = f"""
name: {ENRICHER_NAME}
version: "{ENRICHER_VERSION}"
processor_type: enricher

description: >
  Per-lead demand-gen signal extraction. Fetches LinkedIn posts via
  Saleleads, computes posting stats (3/6/12-month windows on original
  posts only), then runs 5 focused GPT-4o-mini classification prompts to
  derive employment_type / expertise / is_demand_marketer / business_type
  / activity signals / tools. Writes a typed JSONB payload to lead_signals.

  Faithful port of the Acme-AI carousel-1 ETL scripts in
  <your-local-tools-dir>.
  Prompts + truncation rules are byte-for-byte ported so outputs stay
  calibration-compatible with the legacy 388+850 rows already in
  lead_signals @legacy-2026-05.

inputs:
  fields_consulted:
    - leads.linkedin_url
    - leads.headline, leads.bio, leads.current_title, leads.current_company,
      leads.current_company_url  (passed into the GPT prompts as profile context)
    - Saleleads /api/v1/user/posts  (fetches last ~50 posts)
    - OpenAI gpt-4o-mini  (5 chat-completion calls per lead)

outputs:
  writes_to_tables: [lead_signals]
  payload_keys:
    # From compute_stats (no GPT)
    - total_posts_fetched, original_posts_fetched
    - posts_3mo, avg_reactions_3mo
    - posts_6mo, avg_reactions_6mo
    - posts_12mo, avg_reactions_12mo
    # From base prompt
    - business_type, google_ads_posts_6mo, linkedin_ads_posts_6mo
    - conversion_rate_mentions_6mo, collaborations_6mo, client_case_studies_6mo
    - shares_tactics, tools_mentioned
    # From fcmo prompt
    - is_fcmo_or_consultant, employment_type, classification_reasoning
    # From expertise prompt
    - expertise
    # From demand_marketer prompt
    - is_demand_marketer
    # From tools prompt (overwrites base's tools_mentioned)
    - tools_mentioned, primary_tool_category

decision_rule: |
  N/A — this is an enricher (produces facts), not a classifier (makes
  decisions). Downstream classifiers like demand_gen_persona_classifier@1.0
  read these signals via `from-signal:demand_gen_signals_enricher:<key>`
  to apply rules.

rule_changes: |
  1.1 (2026-06-05): two prompt rewrites (FCMO + is_demand_marketer). Same
  model (gpt-4o-mini), same truncation rules, same payload schema — only
  the system prompts change. Calibrated against an 18-lead sample (8 known
  false-negatives, 5 known-correct, 5 known-no) + a 17-lead regression
  test. See FCMO_PROMPT_SYSTEM_V1_2 and DEMAND_MARKETER_PROMPT_SYSTEM_V1_2
  in demand_gen_prompts.py for the new text.

  Rationale:
  - v1.0 silently defaulted to employment_type='unrecognized' for any lead
    where the LinkedIn /user/posts fetch returned empty. 93% of URN-encoded
    engager URLs (/in/ACoAA…) hit that path because Saleleads /user/posts
    doesn't resolve URN-encoded usernames. Result: ~89% of enriched leads
    landed in type=null and were invisible to the persona classifier.
  - v1.0 also marked obvious B2B demand-gen pros as is_demand_marketer='no'
    or 'unclear' (Andrei Zinkevich, Phil Gamache, Sara Stella Lattanzio,
    Carrie Mott, dr. Alice Teodorescu, Pete Vomocil, …) because the old
    prompt required "hands-on channel execution" and explicitly excluded
    "thought leaders posting about marketing trends".

  v1.1 changes:
  - FCMO prompt: classify from headline+title+company when posts are
    absent. When title/company are blank, parse "Founder of X / CEO @ X /
    Fractional CMO" patterns from the headline as evidence. Only return
    'unrecognized' when even the headline is vague.
  - is_demand_marketer prompt: broader definition — includes practitioners,
    leaders, advisors, agency founders, AND thought leaders. Hands-on
    channel execution NOT required. Explicit NO rule for outbound-as-a-
    service / SDR-agency / cold-call-automation / appointment-setting
    (these are sales services, not marketing).

  Calibration outcomes (vs v1.0 on the test sample):
  - FIX_ME cohort (false-negatives): 8/8 corrected, 5/5 fresh holdouts also corrected.
  - STAY_NO cohort (sales / CS / B2C): 5/5 held the line.
  - STAY_CORRECT cohort: minor drift on edge cases that v1.0 had wrong
    (e.g. Maren Hogan, B2B marketing agency CEO, v1.0 said dm:no → v1.1 yes).
  - Two earlier v1.1 regressions (Paul Denham, Maya Kaufman) fixed in v1.2.

  1.0 (2026-05-27): first executable version. Replaces the one-off
  Node.js ETL (analyze-cmo-profiles.js, repass-fcmo.js, repass-expertise.js,
  repass-demand-marketer.js, repass-tools.js, merge-cmo-analysis.js).
  Same model (gpt-4o-mini), same prompts, same truncation rules.
  Difference from legacy ETL: no on-disk CSV/jsonl artifacts — writes
  directly to lead_signals.
"""


# ── Saleleads HTTP ────────────────────────────────────────────────────────

def _api_key() -> str:
    return os.environ.get("FRESH_LINKEDIN_DATA_API_KEY", "")


def saleleads_get(path: str, params: dict, retries: int = 6,
                  backoff_base: int = 15) -> dict | None:
    """GET with retry-on-429. Returns None on hard failure.
    Mirrors engagers_research.saleleads_get."""
    qs = urllib.parse.urlencode(params)
    url = f"{SL_BASE}{path}?{qs}"
    headers = {
        "x-rapidapi-host": SL_HOST,
        "x-rapidapi-key": _api_key(),
        "User-Agent": "curl/7.88.1",
        "Accept": "application/json",
    }
    last_err = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=headers, method="GET")
            with urllib.request.urlopen(req, timeout=45) as r:
                d = json.loads(r.read())
        except urllib.error.HTTPError as e:
            last_err = f"HTTP {e.code}"
            try: body = e.read().decode()[:200]
            except: body = ""
            if e.code == 429 or 500 <= e.code < 600:
                time.sleep(backoff_base * (attempt + 1))
                continue
            print(f"    ! HTTP {e.code} {path}: {body}")
            return None
        except Exception as e:
            last_err = e
            time.sleep(backoff_base * (attempt + 1))
            continue
        if d.get("success") is True:
            return d
        msg = (d.get("message") or "").lower()
        if "429" in msg or "rate" in msg or "denied" in msg or "undefined" in msg:
            time.sleep(backoff_base * (attempt + 1))
            last_err = msg
            continue
        return d
    print(f"    ! saleleads_get FAILED after {retries} retries: {last_err}")
    return None


_URN_LIKE_RE = re.compile(r"^[a-z0-9_-]{20,}$", re.I)


def _slug_from_url(url: str) -> str:
    m = re.search(r"linkedin\.com/in/([^/?#]+)", (url or "").lower())
    return m.group(1).rstrip("/") if m else ""


def _looks_urn_slug(s: str) -> bool:
    if not s: return False
    return bool(_URN_LIKE_RE.match(s)) and s.startswith(("acoaa", "acwaa", "acuaa", "acaaa"))


def fetch_user_posts(linkedin_url: str, max_pages: int = 3) -> list[dict]:
    """Saleleads /api/v1/user/posts — returns the raw posts list.
    20 posts per page; 3 pages = up to 60 posts, matches the ~50 the
    legacy ETL fetched.

    Saleleads' `username` parameter is polymorphic — accepts BOTH vanity
    slugs (e.g. `ofer-miller`) AND URN-encoded forms (e.g.
    `ACwAAApPi4EBlRT4TVtWIiCkrSDZe3dq5PQ2_8w`). We pass the slug as-is
    without trying to "resolve" URN→vanity first; Saleleads handles it.
    """
    slug = _slug_from_url(linkedin_url)
    if not slug:
        return []
    all_posts: list[dict] = []
    for page in range(1, max_pages + 1):
        d = saleleads_get("/api/v1/user/posts", {"username": slug, "page": page})
        if not d or not d.get("success"):
            break
        data = d.get("data") or []
        if not data:
            break
        all_posts.extend(data)
        # Saleleads returns ~20 per page; only stop when a page is empty.
        if len(data) < 20:
            break
    return all_posts


# ── OpenAI call (one prompt, one lead) ────────────────────────────────────

def _openai_client():
    try:
        from openai import OpenAI
    except ImportError:
        sys.exit("openai package required. pip install openai")
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        sys.exit("OPENAI_API_KEY env var required.")
    return OpenAI(api_key=api_key)


def run_prompt(client, prompt_name: str, profile: dict, posts: list[dict]) -> dict:
    """Run one of the 5 prompts and return the parsed JSON."""
    spec = PROMPTS[prompt_name]
    user_text = spec["user_builder"](profile, posts)
    resp = client.chat.completions.create(
        model=MODEL,
        temperature=TEMPERATURE,
        max_tokens=spec["max_tokens"],
        response_format=RESPONSE_FORMAT,
        messages=[
            {"role": "system", "content": spec["system"]},
            {"role": "user", "content": user_text},
        ],
    )
    raw = resp.choices[0].message.content
    return json.loads(raw)


# ── Per-lead enrichment pipeline ──────────────────────────────────────────

def enrich_lead(client, lead_row: dict, posts: list[dict] | None = None) -> dict:
    """Full pipeline: fetch posts (if not provided) → compute stats →
    run 5 prompts (in parallel) → merge into one payload. Returns the
    payload dict; caller writes to DB."""
    if posts is None:
        posts = fetch_user_posts(lead_row["linkedin_url"])

    # Profile context for prompts. Reuses what's already on the lead row;
    # company_industry / company_description aren't on the leads table so
    # left empty (prompts fall through with "(unknown)").
    profile = {
        "name": lead_row.get("name") or "",
        "headline": lead_row.get("headline") or "",
        "bio": lead_row.get("bio") or "",
        "current_title": lead_row.get("current_title") or "",
        "current_company": lead_row.get("current_company") or "",
        "current_company_url": lead_row.get("current_company_url") or "",
        "company_description": "",  # not in leads table
        "company_industry": "",     # not in leads table
        "company_website": lead_row.get("current_company_url") or "",
    }

    # Run all 5 GPT calls in parallel.
    payload: dict = compute_stats(posts)
    with ThreadPoolExecutor(max_workers=5) as ex:
        futures = {
            ex.submit(run_prompt, client, name, profile, posts): name
            for name in PROMPTS
        }
        for f in as_completed(futures):
            name = futures[f]
            try:
                result = f.result()
                # Per the JS merge order: tools prompt's tools_mentioned
                # overrides base's. We process in result order; since base
                # may finish first, we update payload, then tools overwrites.
                # If tools finishes before base, base.tools_mentioned would
                # overwrite tools's. Avoid that:
                if name == "base":
                    # Drop base's tools_mentioned if tools prompt already set it.
                    if "tools_mentioned" in payload:
                        result.pop("tools_mentioned", None)
                payload.update(result)
            except Exception as e:
                print(f"    ! prompt={name} failed: {e}")

    return payload


# ── DB helpers ────────────────────────────────────────────────────────────

def lead_already_has_signal(cur, lead_id) -> bool:
    cur.execute("""
        SELECT 1 FROM lead_signals
        WHERE lead_id=%s AND enricher_name=%s AND enricher_version=%s
    """, (lead_id, ENRICHER_NAME, ENRICHER_VERSION))
    return cur.fetchone() is not None


def upsert_signal(cur, lead_id, payload: dict, who: str) -> None:
    cur.execute("""
        INSERT INTO lead_signals
          (lead_id, enricher_name, enricher_version, payload, enriched_at, enriched_by)
        VALUES (%s, %s, %s, %s, now(), %s)
        ON CONFLICT (lead_id, enricher_name, enricher_version) DO UPDATE
          SET payload = EXCLUDED.payload,
              enriched_at = now(),
              enriched_by = EXCLUDED.enriched_by
    """, (lead_id, ENRICHER_NAME, ENRICHER_VERSION, Jsonb(payload), who))


def resolve_targets(cur, *, lead_url: str | None, lead_file: str | None,
                    where_tag: str | None) -> list[dict]:
    """Returns [{id, linkedin_url, name, headline, current_title,
    current_company, current_company_url, bio}]."""
    base_sql = """
        SELECT id, linkedin_url, name, headline, current_title,
               current_company, current_company_url, bio
        FROM leads
    """
    if lead_url:
        cur.execute(base_sql + " WHERE linkedin_url = %s",
                    (normalize_linkedin_url(lead_url),))
    elif lead_file:
        urls = _read_urls_from_file(Path(lead_file))
        if not urls: return []
        cur.execute(base_sql + " WHERE linkedin_url = ANY(%s)", (urls,))
    elif where_tag:
        cur.execute(base_sql + """
            WHERE id IN (SELECT lead_id FROM lead_tags WHERE tag = %s)
        """, (where_tag,))
    else:
        return []
    cols = ["id","linkedin_url","name","headline","current_title",
            "current_company","current_company_url","bio"]
    return [dict(zip(cols, r)) for r in cur.fetchall()]


def _read_urls_from_file(path: Path) -> list[str]:
    if path.suffix.lower() in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("openpyxl required for XLSX. pip install openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h) if h is not None else "" for h in rows[0]]
        url_col = next((i for i,h in enumerate(headers)
                        if h in ("linkedin_url","profile_url","url","LinkedIn URL")), None)
        if url_col is None: sys.exit("No URL column in XLSX")
        return [normalize_linkedin_url(str(r[url_col])) for r in rows[1:] if r[url_col]]
    with open(path, newline="") as f:
        reader = csv.DictReader(f)
        col = next((c for c in ("linkedin_url","profile_url","url","LinkedIn URL")
                    if c in (reader.fieldnames or [])), None)
        if col is None: sys.exit("No URL column in CSV")
        return [normalize_linkedin_url(row[col]) for row in reader if row.get(col)]


# ── Main ──────────────────────────────────────────────────────────────────

def main() -> int:
    parser = argparse.ArgumentParser(
        description="demand_gen_signals_enricher@1.0 — fetch+classify lead signals into lead_signals.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--client", required=True)
    target = parser.add_mutually_exclusive_group(required=True)
    target.add_argument("--lead-url", help="A single LinkedIn URL.")
    target.add_argument("--lead-file", help="CSV/XLSX with a URL column.")
    target.add_argument("--where-tag", help="Every lead carrying this tag.")
    parser.add_argument("--refresh", action="store_true",
                        help="Re-enrich even if v1.0 signals already exist.")
    parser.add_argument("--limit", type=int, default=0,
                        help="Cap number of leads enriched per run (0=unlimited).")
    parser.add_argument("--dry-run", action="store_true",
                        help="Run the pipeline but don't write to lead_signals.")
    args = parser.parse_args()

    # Self-register in processors registry
    try:
        _, _, action = register_processor_from_yaml(
            args.client, YAML_SPEC, created_by="demand-gen-signals-enricher")
        if action == "inserted":
            print(f"✓ registered processor {ENRICHER_NAME}@{ENRICHER_VERSION}")
    except Exception as e:
        print(f"⚠ self-registration failed (continuing): {e}", file=sys.stderr)

    client = _openai_client()

    with connect(args.client) as conn:
        with conn.cursor() as cur:
            targets = resolve_targets(
                cur, lead_url=args.lead_url, lead_file=args.lead_file,
                where_tag=args.where_tag)
        if not targets:
            print("No target leads.")
            return 0

        if args.limit:
            targets = targets[:args.limit]

        print(f"Targets: {len(targets)} lead(s)")
        if args.dry_run:
            print("(dry-run — no writes)")

        done = skipped = errors = 0
        for i, lead in enumerate(targets, 1):
            with conn.cursor() as cur:
                if not args.refresh and lead_already_has_signal(cur, lead["id"]):
                    skipped += 1
                    print(f"[{i}/{len(targets)}] ⊝ skip (has signals): {lead['name'] or lead['linkedin_url']}")
                    continue
            try:
                payload = enrich_lead(client, lead)
                if args.dry_run:
                    print(f"[{i}/{len(targets)}] (dry-run) {lead['name']}: posts_3mo={payload.get('posts_3mo')} "
                          f"emp={payload.get('employment_type')!r} expertise={payload.get('expertise')!r} "
                          f"dmkt={payload.get('is_demand_marketer')!r}")
                else:
                    with conn.cursor() as cur:
                        upsert_signal(cur, lead["id"], payload,
                                      who="demand-gen-signals-enricher-1.0")
                    conn.commit()
                    print(f"[{i}/{len(targets)}] ✓ {lead['name']}: posts_3mo={payload.get('posts_3mo')} "
                          f"emp={payload.get('employment_type')!r} expertise={payload.get('expertise')!r} "
                          f"dmkt={payload.get('is_demand_marketer')!r}")
                done += 1
            except Exception as e:
                errors += 1
                print(f"[{i}/{len(targets)}] ✗ {lead['name'] or lead['linkedin_url']}: {e}")

        print(f"\nSummary: enriched={done}  skipped={skipped}  errors={errors}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
